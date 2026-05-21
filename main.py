import os
import re
import tempfile
import sys
import json
import logging
import subprocess
import hashlib
import threading
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timedelta

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

import gspread
from oauth2client.service_account import ServiceAccountCredentials

from PIL import Image, ImageDraw, ImageFont

SPREADSHEET_ID = "1soHrN7Iqd3jk9iLGdUGK9APxVfRBwWXHxoI8x2Hsh1o"
SHEET_NAME = "data"
APP_NAME = "TakSklad"
APP_EXECUTABLE_NAME = "TakSklad.exe"

def get_app_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

APP_DIR = get_app_dir()
CREDENTIALS_FILE = os.path.join(APP_DIR, "credentials.json")
LOG_FILE = os.path.join(APP_DIR, "TakSklad.log")
BACKUP_DIR = os.path.join(APP_DIR, "scan_backups")
REPORTS_DIR = os.path.join(APP_DIR, "reports")
PENDING_PRINTS_FILE = os.path.join(APP_DIR, "pending_prints.json")
PENDING_SAVES_FILE = os.path.join(APP_DIR, "pending_saves.json")
PENDING_TELEGRAM_FILE = os.path.join(APP_DIR, "pending_telegram.json")
TELEGRAM_STATE_FILE = os.path.join(APP_DIR, "telegram_state.json")
PRINT_SETTINGS_FILE = os.path.join(APP_DIR, "print_settings.json")
PRODUCT_CATALOG_FILE = os.path.join(APP_DIR, "product_catalog.json")
IMPORT_HISTORY_FILE = os.path.join(APP_DIR, "import_history.json")
TELEGRAM_SETTINGS_FILE = os.path.join(APP_DIR, "telegram_settings.json")
YANDEX_GEOCODER_KEY_FILE = os.path.join(APP_DIR, "yandex_geocoder_key.txt")
YANDEX_GEOCODER_API_KEY = "7c455cc8-0cda-46da-ac5c-e32297c2fec0"
APP_VERSION = "1.1.3"
UPDATE_INFO_URL = os.environ.get(
    "PKIS_UPDATE_INFO_URL",
    "https://raw.githubusercontent.com/1fear/pKIS/main/version.json",
).strip()
UPDATE_CHECK_TIMEOUT_SECONDS = 8
UPDATE_DOWNLOAD_TIMEOUT_SECONDS = 120
TELEGRAM_FILE_DOWNLOAD_TIMEOUT_SECONDS = 120
EXCEL_IMPORT_EXTENSIONS = {".xlsx", ".xlsm"}

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    encoding="utf-8"
)

REQUIRED_COLUMNS = [
    "Дата получения заказа",
    "Тип оплаты",
    "Клиент",
    "Адрес",
    "Торговый представитель",
    "Товары",
    "Кол-во ШТ",
    "Кол-во блок",
    "Отсканированные коды",
]

STATUS_COLUMN = "Статус"
STATUS_NOT_COMPLETED = "Не выполнено"
STATUS_COMPLETED = "Выполнено"

WORKING_COLUMNS = REQUIRED_COLUMNS + [STATUS_COLUMN]

SERVICE_COLUMNS = [
    "ID заказа",
    "ID импорта",
    "Источник файла",
    "Строка файла",
    "Дата импорта",
]

SERVICE_COLUMN_START_INDEX = 26  # AA, zero-based

SOURCE_REQUIRED_ALIASES = {
    "client": ["ФИО или Наименование торговой точки", "Клиент", "Юр. лицо", "Юр лицо", "Наименование"],
    "payment": ["Тип оплаты", "Оплата"],
    "product": ["Наименование Товара", "Товары", "Товар", "Номенклатура"],
    "quantity": ["Кол-во", "Количество", "Кол-во ШТ", "Количество ШТ"],
}

SOURCE_OPTIONAL_ALIASES = {
    "date": ["Дата доставки", "Дата получения заказа", "Дата заказа", "Дата"],
    "address": ["Адрес доставки", "Адрес"],
    "coords": ["Координаты", "Координаты доставки"],
    "representative": ["Торговый представитель", "ТП", "Менеджер", "Номер телефона"],
}

DEFAULT_PIECES_PER_BLOCK = 10

LABEL_WIDTH_MM = 100
LABEL_HEIGHT_MM = 100
LABEL_DPI = 203
KIZ_MIN_LENGTH = 20
KIZ_MAX_LENGTH = 120

BG_MAIN = "#f5f7fa"
BG_CARD = "#ffffff"
FG_TEXT = "#1a1f2e"
FG_MUTED = "#6b7280"
ACCENT = "#4f46e5"
SUCCESS = "#10b981"
INFO = "#3b82f6"
WARNING = "#f59e0b"
DANGER = "#ef4444"
ERROR_BG = "#fee2e2"
ERROR_FG = "#dc2626"
BORDER = "#e5e7eb"
DISABLED_BG = "#e5e7eb"
DISABLED_FG = "#94a3b8"

def darken_hex(color, factor=0.9):
    color = color.lstrip("#")
    if len(color) != 6:
        return "#" + color
    try:
        red = int(color[0:2], 16)
        green = int(color[2:4], 16)
        blue = int(color[4:6], 16)
    except ValueError:
        return "#" + color
    return "#{:02x}{:02x}{:02x}".format(
        max(0, int(red * factor)),
        max(0, int(green * factor)),
        max(0, int(blue * factor)),
    )

class AppButton(tk.Frame):
    def __init__(
        self,
        parent,
        text="",
        bg=ACCENT,
        fg="white",
        font=("Segoe UI", 10, "bold"),
        command=None,
        state="normal",
        padx=14,
        pady=8,
        cursor="hand2",
        disabled_bg=DISABLED_BG,
        disabled_fg=DISABLED_FG,
        **kwargs
    ):
        frame_kwargs = {
            "bg": bg,
            "bd": 0,
            "highlightthickness": 0,
            "takefocus": 0,
        }
        for key in ("width", "height"):
            if key in kwargs:
                frame_kwargs[key] = kwargs[key]
        super().__init__(parent, **frame_kwargs)
        self._text = text
        self._normal_bg = bg
        self._normal_fg = fg
        self._active_bg = darken_hex(bg, 0.92)
        self._disabled_bg = disabled_bg
        self._disabled_fg = disabled_fg
        self._command = command
        self._state = state
        self._cursor = cursor

        self.label = tk.Label(
            self,
            text=text,
            bg=bg,
            fg=fg,
            font=font,
            bd=0,
            padx=padx,
            pady=pady,
            anchor="center"
        )
        self.label.pack(fill="both", expand=True)

        for widget in (self, self.label):
            widget.bind("<Button-1>", self._on_click)
            widget.bind("<Enter>", self._on_enter)
            widget.bind("<Leave>", self._on_leave)

        self.config(state=state)

    def _on_click(self, _event=None):
        if self._state != "normal":
            return "break"
        if self._command:
            self._command()
        return "break"

    def _on_enter(self, _event=None):
        if self._state == "normal":
            self._paint(self._active_bg, self._normal_fg)

    def _on_leave(self, _event=None):
        self._refresh_style()

    def _paint(self, bg, fg):
        tk.Frame.configure(self, bg=bg)
        self.label.configure(bg=bg, fg=fg)

    def _refresh_style(self):
        if self._state == "normal":
            self._paint(self._normal_bg, self._normal_fg)
            tk.Frame.configure(self, cursor=self._cursor)
            self.label.configure(cursor=self._cursor)
        else:
            self._paint(self._disabled_bg, self._disabled_fg)
            tk.Frame.configure(self, cursor="")
            self.label.configure(cursor="")

    def configure(self, cnf=None, **kwargs):
        options = {}
        if cnf:
            options.update(cnf)
        options.update(kwargs)

        if "state" in options:
            self._state = options.pop("state")
        if "text" in options:
            self._text = options.pop("text")
            self.label.configure(text=self._text)
        if "command" in options:
            self._command = options.pop("command")
        if "bg" in options:
            self._normal_bg = options.pop("bg")
            self._active_bg = darken_hex(self._normal_bg, 0.92)
        if "background" in options:
            self._normal_bg = options.pop("background")
            self._active_bg = darken_hex(self._normal_bg, 0.92)
        if "fg" in options:
            self._normal_fg = options.pop("fg")
        if "foreground" in options:
            self._normal_fg = options.pop("foreground")
        if "font" in options:
            self.label.configure(font=options.pop("font"))
        if "cursor" in options:
            self._cursor = options.pop("cursor")

        ignored = {
            "relief", "activebackground", "activeforeground",
            "selectbackground", "selectforeground", "disabledforeground",
        }
        for key in list(options.keys()):
            if key in ignored:
                options.pop(key)

        if options:
            tk.Frame.configure(self, **options)
        self._refresh_style()

    config = configure

    def cget(self, key):
        if key == "state":
            return self._state
        if key == "text":
            return self._text
        return tk.Frame.cget(self, key)

def format_exception_message(title, exc):
    return (
        f"{title}\n\n"
        f"Причина: {exc}\n\n"
        f"Подробности записаны в лог:\n{LOG_FILE}"
    )

def show_exception_message(title, exc):
    logging.exception(title)
    try:
        messagebox.showerror("Ошибка", format_exception_message(title, exc))
    except Exception:
        pass

def global_exception_handler(exc_type, exc_value, exc_traceback):
    if issubclass(exc_type, KeyboardInterrupt):
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
    logging.error(
        "Неперехваченная ошибка",
        exc_info=(exc_type, exc_value, exc_traceback)
    )
    try:
        messagebox.showerror(
            "Критическая ошибка",
            format_exception_message("Неперехваченная ошибка", exc_value)
        )
    except Exception:
        pass

sys.excepthook = global_exception_handler

def get_google_client():
    scope = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, scope)
    return gspread.authorize(creds)

def clean_date_value(date_value):
    if date_value is None:
        return None
    date_str = str(date_value).strip()
    date_str = re.sub(r'^[\'"]+|[\'"]+$', '', date_str)
    if ' ' in date_str:
        date_str = date_str.split()[0]
    return date_str

def parse_date_to_standard(date_value):
    cleaned = clean_date_value(date_value)
    if not cleaned:
        return None
    formats = ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d.%m.%y", "%Y.%m.%d"]
    for fmt in formats:
        try:
            parsed = datetime.strptime(cleaned, fmt)
            return parsed.strftime("%d.%m.%Y")
        except:
            continue
    return cleaned

def normalize_text(value):
    return str(value or "").strip()

def clean_file_name(file_name, fallback="file"):
    text = normalize_text(file_name).replace("\\", "/")
    name = os.path.basename(text).strip()
    return name or fallback

def is_supported_excel_file_name(file_name):
    extension = os.path.splitext(clean_file_name(file_name))[1].lower()
    return extension in EXCEL_IMPORT_EXTENSIONS

def normalize_header_name(value):
    return normalize_text(value).replace("\ufeff", "")

def get_header_index(header):
    return {normalize_header_name(col): idx for idx, col in enumerate(header) if normalize_header_name(col)}

def get_header_indices(header, column_name):
    normalized_column = normalize_header_name(column_name)
    return [
        idx
        for idx, col in enumerate(header)
        if normalize_header_name(col) == normalized_column
    ]

def column_index_to_letter(index):
    index += 1
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters

def get_cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    return normalize_text(row[idx])

def split_codes(codes_str):
    if not codes_str:
        return []
    codes = []
    for line in str(codes_str).splitlines():
        code = line.strip()
        if code:
            codes.append(code)
    return codes

def normalize_payment_type(value):
    payment = normalize_text(value).lower().replace("ё", "е")
    if "терминал" in payment:
        return "terminal"
    if "перечис" in payment or "безнал" in payment:
        return "transfer"
    return "unknown"

def parse_int_value(value):
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    value_str = normalize_text(value).replace(" ", "").replace(",", ".")
    if not value_str:
        return 0
    try:
        return int(float(value_str))
    except ValueError:
        return 0

def load_json_file(path, default):
    try:
        if not os.path.exists(path):
            return default
        with open(path, "r", encoding="utf-8") as json_file:
            data = json.load(json_file)
        return data if data is not None else default
    except Exception:
        logging.exception("Не удалось загрузить JSON-файл: %s", path)
        return default

def save_json_file(path, data):
    try:
        with open(path, "w", encoding="utf-8") as json_file:
            json.dump(data, json_file, ensure_ascii=False, indent=2)
        return True
    except Exception:
        logging.exception("Не удалось сохранить JSON-файл: %s", path)
        return False

def normalize_lookup_text(value):
    text = normalize_text(value).lower().replace("ё", "е")
    text = text.replace("\ufeff", "")
    text = re.sub(r"[*:]+", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()

def file_sha1(path):
    sha1 = hashlib.sha1()
    with open(path, "rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            sha1.update(chunk)
    return sha1.hexdigest()

def make_hash(payload):
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()

def load_yandex_geocoder_key():
    env_key = normalize_text(os.environ.get("YANDEX_GEOCODER_API_KEY"))
    if env_key:
        return env_key
    try:
        if os.path.exists(YANDEX_GEOCODER_KEY_FILE):
            with open(YANDEX_GEOCODER_KEY_FILE, "r", encoding="utf-8") as key_file:
                file_key = normalize_text(key_file.read())
            if file_key:
                return file_key
    except Exception:
        logging.exception("Не удалось прочитать ключ Яндекс Геокодера")
    return normalize_text(YANDEX_GEOCODER_API_KEY)

def normalize_coordinates(value):
    text = normalize_text(value)
    if not text:
        return ""

    numbers = re.findall(r"-?\d+(?:[.,]\d+)?", text)
    if len(numbers) < 2:
        return ""

    try:
        first = float(numbers[0].replace(",", "."))
        second = float(numbers[1].replace(",", "."))
    except ValueError:
        return ""

    if abs(first) <= 90 and abs(second) <= 180:
        lat, lon = first, second
    elif abs(second) <= 90 and abs(first) <= 180:
        lat, lon = second, first
    else:
        return ""

    return f"{lat:.8f},{lon:.8f}".rstrip("0").rstrip(".")

def reverse_geocode_yandex(coords, cache=None):
    import urllib.error
    import urllib.parse
    import urllib.request

    normalized_coords = normalize_coordinates(coords)
    if not normalized_coords:
        return None, "некорректные координаты"

    if cache is not None and normalized_coords in cache:
        return cache[normalized_coords]

    api_key = load_yandex_geocoder_key()
    if not api_key:
        result = (None, "не указан ключ Яндекс Геокодера")
        if cache is not None:
            cache[normalized_coords] = result
        return result

    params = {
        "apikey": api_key,
        "geocode": normalized_coords,
        "format": "json",
        "lang": "ru_RU",
        "sco": "latlong",
        "results": "1",
        "kind": "house",
    }
    url = "https://geocode-maps.yandex.ru/v1/?" + urllib.parse.urlencode(params)

    try:
        with urllib.request.urlopen(url, timeout=15) as response:
            data = json.load(response)
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", "replace")[:300]
        result = (None, f"HTTP {exc.code}: {body}")
        if cache is not None:
            cache[normalized_coords] = result
        return result
    except Exception as exc:
        result = (None, str(exc))
        if cache is not None:
            cache[normalized_coords] = result
        return result

    members = data.get("response", {}).get("GeoObjectCollection", {}).get("featureMember", [])
    if not members:
        result = (None, "адрес не найден")
        if cache is not None:
            cache[normalized_coords] = result
        return result

    obj = members[0].get("GeoObject", {})
    meta = obj.get("metaDataProperty", {}).get("GeocoderMetaData", {})
    address = normalize_text(meta.get("text") or obj.get("name"))
    if not address:
        result = (None, "пустой адрес в ответе Яндекса")
    else:
        result = (address, "")

    if cache is not None:
        cache[normalized_coords] = result
    return result

def load_product_catalog():
    catalog = load_json_file(PRODUCT_CATALOG_FILE, {})
    return catalog if isinstance(catalog, dict) else {}

def save_product_catalog(catalog):
    return save_json_file(PRODUCT_CATALOG_FILE, catalog)

def product_catalog_key(product_name):
    return normalize_lookup_text(product_name)

def get_product_rule(product_name, catalog=None, create=False):
    catalog = catalog if catalog is not None else load_product_catalog()
    key = product_catalog_key(product_name)
    if not key:
        return {
            "name": "",
            "pieces_per_block": DEFAULT_PIECES_PER_BLOCK,
            "requires_kiz": True,
        }
    if key not in catalog and create:
        catalog[key] = {
            "name": normalize_text(product_name),
            "pieces_per_block": DEFAULT_PIECES_PER_BLOCK,
            "requires_kiz": True,
        }
    rule = catalog.get(key, {})
    pieces = parse_int_value(rule.get("pieces_per_block")) or DEFAULT_PIECES_PER_BLOCK
    return {
        "name": rule.get("name") or normalize_text(product_name),
        "pieces_per_block": max(1, pieces),
        "requires_kiz": bool(rule.get("requires_kiz", True)),
    }

def calculate_blocks(quantity, product_name, catalog, warnings=None):
    qty = parse_int_value(quantity)
    rule = get_product_rule(product_name, catalog=catalog, create=True)
    pieces_per_block = rule["pieces_per_block"]
    blocks = (qty + pieces_per_block - 1) // pieces_per_block if qty > 0 else 0
    if warnings is not None and qty > 0 and qty % pieces_per_block != 0:
        warnings.append(
            f"'{product_name}': количество {qty} не делится на {pieces_per_block}, "
            f"план округлён до {blocks} блок."
        )
    return blocks, pieces_per_block

def get_source_header_index(header):
    return {normalize_lookup_text(col): idx for idx, col in enumerate(header) if normalize_lookup_text(col)}

def find_source_column(header_idx, aliases):
    for alias in aliases:
        key = normalize_lookup_text(alias)
        if key in header_idx:
            return header_idx[key]
    return None

def get_source_cell(row, idx):
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    return normalize_text(value)

def make_order_id(record):
    return make_hash({
        "date": parse_date_to_standard(record.get("Дата получения заказа")),
        "payment": normalize_lookup_text(record.get("Тип оплаты")),
        "client": normalize_lookup_text(record.get("Клиент")),
        "address": normalize_lookup_text(record.get("Адрес")),
        "representative": normalize_lookup_text(record.get("Торговый представитель")),
        "product": normalize_lookup_text(record.get("Товары")),
        "quantity": parse_int_value(record.get("Кол-во ШТ")),
        "blocks": parse_int_value(record.get("Кол-во блок")),
    })

def get_plan_blocks(order):
    plan_blocks = parse_int_value(order.get("Кол-во блок", 0))
    if plan_blocks == 0:
        plan_blocks = parse_int_value(order.get("Кол-во блоков", 0))
    return plan_blocks

def is_order_completed(order):
    plan_blocks = get_plan_blocks(order)
    scanned_count = len(split_codes(order.get("Отсканированные коды")))
    return plan_blocks > 0 and scanned_count >= plan_blocks

def is_completed_status(value):
    status = normalize_lookup_text(value)
    if not status:
        return False
    not_completed_markers = [
        "не выполн",
        "невыполн",
        "не готов",
        "неготов",
        "нет",
        "false",
        "0",
    ]
    if any(marker in status for marker in not_completed_markers):
        return False
    completed_markers = [
        "выполн",
        "готов",
        "done",
        "complete",
        "completed",
        "yes",
        "true",
        "1",
    ]
    return any(marker in status for marker in completed_markers)

def get_order_status(order):
    return STATUS_COMPLETED if is_order_completed(order) else STATUS_NOT_COMPLETED

def is_order_active(order):
    status = normalize_text(order.get(STATUS_COLUMN))
    if status:
        return not is_completed_status(status)
    return not is_order_completed(order)

def order_group_key(order):
    client = normalize_text(order.get("Клиент")) or "Клиент не указан"
    payment_type = normalize_text(order.get("Тип оплаты")) or "Оплата не указана"
    address = normalize_text(order.get("Адрес")) or "Адрес не указан"
    return (
        client,
        payment_type,
        address,
    )

def row_matches_order(row, header_idx, order):
    order_id = normalize_text(order.get("ID заказа"))
    order_id_idx = header_idx.get("ID заказа")
    if order_id and order_id_idx is not None and get_cell(row, order_id_idx) == order_id:
        return True

    checks = [
        ("Дата получения заказа", parse_date_to_standard(get_cell(row, header_idx.get("Дата получения заказа"))) == parse_date_to_standard(order.get("Дата получения заказа"))),
        ("Тип оплаты", get_cell(row, header_idx.get("Тип оплаты")) == normalize_text(order.get("Тип оплаты"))),
        ("Клиент", get_cell(row, header_idx.get("Клиент")) == normalize_text(order.get("Клиент"))),
        ("Адрес", get_cell(row, header_idx.get("Адрес")) == normalize_text(order.get("Адрес"))),
        ("Товары", get_cell(row, header_idx.get("Товары")) == normalize_text(order.get("Товары"))),
    ]
    return all(result for _, result in checks)

def validate_sheet_header(header):
    header_idx = get_header_index(header)
    missing = [col for col in REQUIRED_COLUMNS if col not in header_idx]
    return header_idx, missing

def ensure_sheet_columns(sheet, columns):
    all_rows = sheet.get_all_values()
    if not all_rows:
        header = list(columns)
        sheet.append_row(header, value_input_option="USER_ENTERED")
        return header

    header = [normalize_header_name(col) for col in all_rows[0]]
    header_idx = get_header_index(header)
    for column in columns:
        if column not in header_idx:
            header.append(column)
            sheet.update_cell(1, len(header), column)
            header_idx[column] = len(header) - 1
    return header

def build_import_sheet_header():
    header = [""] * (SERVICE_COLUMN_START_INDEX + len(SERVICE_COLUMNS))
    for idx, column in enumerate(WORKING_COLUMNS):
        header[idx] = column
    for offset, column in enumerate(SERVICE_COLUMNS):
        header[SERVICE_COLUMN_START_INDEX + offset] = column
    return header

def get_import_column_targets():
    targets = []
    targets.extend((idx, column) for idx, column in enumerate(WORKING_COLUMNS))
    targets.extend(
        (SERVICE_COLUMN_START_INDEX + offset, column)
        for offset, column in enumerate(SERVICE_COLUMNS)
    )
    return targets

def ensure_import_sheet_columns(sheet):
    all_rows = sheet.get_all_values()
    required_len = SERVICE_COLUMN_START_INDEX + len(SERVICE_COLUMNS)
    if not all_rows:
        header = build_import_sheet_header()
        sheet.append_row(header, value_input_option="USER_ENTERED")
        return header

    header = [normalize_header_name(col) for col in all_rows[0]]
    if len(header) < required_len:
        header.extend([""] * (required_len - len(header)))

    # The import sheet has a fixed layout for warehouse work:
    # A:J are operational fields, AA:AE are technical import fields.
    # If row 1 is empty or incomplete, rewrite only the headers in those slots.
    for target_idx, column in get_import_column_targets():
        header[target_idx] = column

    last_col = column_index_to_letter(len(header) - 1)
    sheet.batch_update([{
        "range": f"A1:{last_col}1",
        "values": [header],
    }], value_input_option="USER_ENTERED")

    return header

def migrate_legacy_service_columns(sheet):
    all_rows = sheet.get_all_values()
    if len(all_rows) <= 1:
        return

    header = [normalize_header_name(col) for col in all_rows[0]]
    updates = []
    clear_ranges = []

    for offset, column in enumerate(SERVICE_COLUMNS):
        target_idx = SERVICE_COLUMN_START_INDEX + offset
        target_has_data = any(get_cell(row, target_idx) for row in all_rows[1:])
        if target_has_data:
            continue

        for source_idx in get_header_indices(header, column):
            if source_idx == target_idx:
                continue
            source_has_data = any(get_cell(row, source_idx) for row in all_rows[1:])
            if not source_has_data:
                continue

            target_col = column_index_to_letter(target_idx)
            source_col = column_index_to_letter(source_idx)
            updates.append({
                "range": f"{target_col}2:{target_col}{len(all_rows)}",
                "values": [[get_cell(row, source_idx)] for row in all_rows[1:]],
            })
            clear_ranges.append(f"{source_col}2:{source_col}{len(all_rows)}")
            break

    if updates:
        sheet.batch_update(updates, value_input_option="USER_ENTERED")
    if clear_ranges:
        sheet.batch_clear(clear_ranges)

def build_import_record_row(record):
    row = [""] * (SERVICE_COLUMN_START_INDEX + len(SERVICE_COLUMNS))
    for idx, column in enumerate(WORKING_COLUMNS):
        row[idx] = record.get(column, "")
    for offset, column in enumerate(SERVICE_COLUMNS):
        row[SERVICE_COLUMN_START_INDEX + offset] = record.get(column, "")
    return row

def ensure_import_sheet_layout(sheet):
    header = ensure_import_sheet_columns(sheet)
    migrate_legacy_service_columns(sheet)
    return header

def get_existing_import_keys(all_rows):
    if not all_rows:
        return set(), set()

    import_indices = get_header_indices(all_rows[0], "ID импорта")
    order_indices = get_header_indices(all_rows[0], "ID заказа")
    import_ids = set()
    order_ids = set()

    for row in all_rows[1:]:
        for import_idx in import_indices:
            import_id = get_cell(row, import_idx)
            if import_id:
                import_ids.add(import_id)
        for order_idx in order_indices:
            order_id = get_cell(row, order_idx)
            if order_id:
                order_ids.add(order_id)

    return import_ids, order_ids

def parse_excel_order_files(file_paths, source_names=None):
    import openpyxl

    catalog = load_product_catalog()
    raw_rows = []
    errors = []
    warnings = []
    source_rows_count = 0
    geocoded_count = 0
    geocode_failed_count = 0
    geocode_cache = {}
    source_names = source_names or {}
    source_names_by_path = {
        os.path.abspath(path): clean_file_name(name, os.path.basename(path))
        for path, name in source_names.items()
    }

    for file_path in file_paths:
        file_name = source_names_by_path.get(os.path.abspath(file_path)) or clean_file_name(os.path.basename(file_path))
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        except Exception as exc:
            errors.append(f"{file_name}: не удалось открыть файл ({exc})")
            continue

        sheet_name = "Заявки" if "Заявки" in workbook.sheetnames else workbook.sheetnames[0]
        worksheet = workbook[sheet_name]
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            errors.append(f"{file_name}: лист пустой")
            continue

        header_idx = get_source_header_index(header)
        columns = {}
        missing = []
        for key, aliases in SOURCE_REQUIRED_ALIASES.items():
            idx = find_source_column(header_idx, aliases)
            if idx is None:
                missing.append(aliases[0])
            columns[key] = idx

        for key, aliases in SOURCE_OPTIONAL_ALIASES.items():
            columns[key] = find_source_column(header_idx, aliases)

        if missing:
            errors.append(f"{file_name}: нет обязательных колонок: {', '.join(missing)}")
            continue

        source_file_hash = file_sha1(file_path)
        source_file_sha256 = file_sha256(file_path)
        for row_number, row in enumerate(rows_iter, start=2):
            if not row or not any(normalize_text(cell) for cell in row if cell is not None):
                continue

            source_rows_count += 1
            client = get_source_cell(row, columns["client"])
            payment = get_source_cell(row, columns["payment"])
            product = get_source_cell(row, columns["product"])
            quantity = parse_int_value(get_source_cell(row, columns["quantity"]))

            if not client or not payment or not product or quantity <= 0:
                warnings.append(f"{file_name}, строка {row_number}: пропущена, не заполнены клиент/оплата/товар/количество")
                continue

            date_value = parse_date_to_standard(get_source_cell(row, columns.get("date"))) or datetime.now().strftime("%d.%m.%Y")
            address = get_source_cell(row, columns.get("address"))
            coords = get_source_cell(row, columns.get("coords"))
            if not address and coords:
                geocoded_address, geocode_error = reverse_geocode_yandex(coords, cache=geocode_cache)
                if geocoded_address:
                    address = geocoded_address
                    geocoded_count += 1
                else:
                    geocode_failed_count += 1
                    address = f"Координаты: {coords}"
                    warnings.append(f"{file_name}, строка {row_number}: адрес по координатам не получен ({geocode_error})")
            if not address:
                warnings.append(f"{file_name}, строка {row_number}: адрес пустой")
                address = "Адрес не указан"

            representative = get_source_cell(row, columns.get("representative"))
            source_id = make_hash({
                "file_hash": source_file_hash,
                "sheet": sheet_name,
                "row": row_number,
            })

            raw_rows.append({
                "date": date_value,
                "payment": payment,
                "client": client,
                "address": address,
                "representative": representative,
                "product": product,
                "quantity": quantity,
                "source_id": source_id,
                "source_file": file_name,
                "source_file_sha256": source_file_sha256,
                "source_row": row_number,
            })

    grouped = {}
    for row in raw_rows:
        key = (
            row["date"],
            normalize_lookup_text(row["payment"]),
            normalize_lookup_text(row["client"]),
            normalize_lookup_text(row["address"]),
            normalize_lookup_text(row["representative"]),
            normalize_lookup_text(row["product"]),
        )
        if key not in grouped:
            grouped[key] = row.copy()
            grouped[key]["source_ids"] = []
            grouped[key]["source_rows"] = []
            grouped[key]["source_files"] = set()
            grouped[key]["source_file_sha256"] = set()
        else:
            grouped[key]["quantity"] += row["quantity"]
        grouped[key]["source_ids"].append(row["source_id"])
        grouped[key]["source_rows"].append(str(row["source_row"]))
        grouped[key]["source_files"].add(row["source_file"])
        grouped[key]["source_file_sha256"].add(row["source_file_sha256"])

    records = []
    for item in grouped.values():
        blocks, pieces_per_block = calculate_blocks(item["quantity"], item["product"], catalog, warnings)
        record = {
            "Дата получения заказа": item["date"],
            "Тип оплаты": item["payment"],
            "Клиент": item["client"],
            "Адрес": item["address"],
            "Торговый представитель": item["representative"],
            "Товары": item["product"],
            "Кол-во ШТ": item["quantity"],
            "Кол-во блок": blocks,
            "Отсканированные коды": "",
            STATUS_COLUMN: STATUS_NOT_COMPLETED,
            "ID импорта": make_hash(sorted(item["source_ids"])),
            "Источник файла": ", ".join(sorted(item["source_files"])),
            "Строка файла": ", ".join(item["source_rows"]),
            "Дата импорта": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
        }
        record["ID заказа"] = make_order_id(record)
        record["_pieces_per_block"] = pieces_per_block
        record["_source_file_sha256"] = sorted(item["source_file_sha256"])
        records.append(record)

    save_product_catalog(catalog)

    return {
        "records": records,
        "errors": errors,
        "warnings": warnings,
        "source_rows_count": source_rows_count,
        "files_count": len(file_paths),
        "geocoded_count": geocoded_count,
        "geocode_failed_count": geocode_failed_count,
    }

def prepare_excel_import(file_paths, source_names=None):
    parsed = parse_excel_order_files(file_paths, source_names=source_names)
    client = get_google_client()
    sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)
    ensure_import_sheet_layout(sheet)
    all_rows = sheet.get_all_values()

    existing_import_ids, existing_order_ids = get_existing_import_keys(all_rows)
    new_records = []
    duplicate_records = []

    for record in parsed["records"]:
        if record.get("ID импорта") in existing_import_ids or record.get("ID заказа") in existing_order_ids:
            duplicate_records.append(record)
        else:
            new_records.append(record)

    parsed["new_records"] = new_records
    parsed["duplicate_records"] = duplicate_records
    parsed["clients_count"] = len({record.get("Клиент") for record in new_records})
    parsed["products_count"] = len({record.get("Товары") for record in new_records})
    parsed["blocks_count"] = sum(parse_int_value(record.get("Кол-во блок")) for record in new_records)
    parsed["quantity_count"] = sum(parse_int_value(record.get("Кол-во ШТ")) for record in new_records)
    return parsed

def extract_record_file_hashes(records):
    hashes = set()
    for record in records:
        raw_hashes = record.get("_source_file_sha256", [])
        if isinstance(raw_hashes, str):
            raw_hashes = [raw_hashes]
        for file_hash in raw_hashes:
            normalized_hash = normalize_text(file_hash).lower()
            if normalized_hash:
                hashes.add(normalized_hash)
    return hashes

def find_successful_import_by_file_hash(file_hash):
    normalized_target = normalize_text(file_hash).lower()
    if not normalized_target:
        return None

    history = load_json_file(IMPORT_HISTORY_FILE, [])
    if not isinstance(history, list):
        return None

    for item in reversed(history):
        if not isinstance(item, dict) or parse_int_value(item.get("imported")) <= 0:
            continue
        entry_hashes = set()
        for key in ("source_file_hashes_sha256", "file_hashes_sha256", "file_sha256"):
            raw_hashes = item.get(key, [])
            if isinstance(raw_hashes, str):
                raw_hashes = [raw_hashes]
            for value in raw_hashes:
                normalized_hash = normalize_text(value).lower()
                if normalized_hash:
                    entry_hashes.add(normalized_hash)
        if normalized_target in entry_hashes:
            return item
    return None

def append_import_records(records):
    if not records:
        return {"imported": 0, "duplicates": 0}

    client = get_google_client()
    sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)
    ensure_import_sheet_layout(sheet)
    all_rows = sheet.get_all_values()
    existing_import_ids, existing_order_ids = get_existing_import_keys(all_rows)

    rows_to_append = []
    appended_records = []
    duplicates = 0
    for record in records:
        if record.get("ID импорта") in existing_import_ids or record.get("ID заказа") in existing_order_ids:
            duplicates += 1
            continue
        rows_to_append.append(build_import_record_row(record))
        appended_records.append(record)
        existing_import_ids.add(record.get("ID импорта"))
        existing_order_ids.add(record.get("ID заказа"))

    if rows_to_append:
        start_row = len(all_rows) + 1
        end_row = start_row + len(rows_to_append) - 1
        sheet.batch_update([{
            "range": f"A{start_row}:AE{end_row}",
            "values": rows_to_append,
        }], value_input_option="USER_ENTERED")

    history = load_json_file(IMPORT_HISTORY_FILE, [])
    if not isinstance(history, list):
        history = []
    history.append({
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "imported": len(rows_to_append),
        "duplicates": duplicates,
        "sources": sorted({record.get("Источник файла", "") for record in records}),
        "source_file_hashes_sha256": sorted(extract_record_file_hashes(appended_records)),
    })
    save_json_file(IMPORT_HISTORY_FILE, history[-200:])

    return {"imported": len(rows_to_append), "duplicates": duplicates}

def get_all_existing_codes(sheet):
    try:
        all_rows = sheet.get_all_values()
        if not all_rows:
            return set()
        header_idx = get_header_index(all_rows[0])
        codes_idx = header_idx.get("Отсканированные коды")
        if codes_idx is None:
            logging.warning("Колонка 'Отсканированные коды' не найдена")
            return set()

        all_codes = set()
        for row in all_rows[1:]:
            for code in split_codes(get_cell(row, codes_idx)):
                all_codes.add(code)
        return all_codes
    except Exception as e:
        logging.exception("Не удалось загрузить существующие коды")
        return set()

def find_code_details_in_rows(all_rows, code):
    if not all_rows:
        return []

    header_idx, missing = validate_sheet_header(all_rows[0])
    if missing:
        raise ValueError("В таблице не найдены обязательные колонки: " + ", ".join(missing))

    codes_idx = header_idx.get("Отсканированные коды")
    details = []
    for row_number, row in enumerate(all_rows[1:], start=2):
        row_codes = split_codes(get_cell(row, codes_idx))
        if code not in row_codes:
            continue

        details.append({
            "row_number": row_number,
            "date": get_cell(row, header_idx.get("Дата получения заказа")),
            "payment": get_cell(row, header_idx.get("Тип оплаты")),
            "client": get_cell(row, header_idx.get("Клиент")),
            "address": get_cell(row, header_idx.get("Адрес")),
            "representative": get_cell(row, header_idx.get("Торговый представитель")),
            "product": get_cell(row, header_idx.get("Товары")),
            "quantity": get_cell(row, header_idx.get("Кол-во ШТ")),
            "blocks": get_cell(row, header_idx.get("Кол-во блок")),
            "status": get_cell(row, header_idx.get(STATUS_COLUMN)),
            "codes_count": len(row_codes),
        })
    return details

def find_code_details_in_sheet(sheet, code):
    if not sheet:
        return []
    return find_code_details_in_rows(sheet.get_all_values(), code)

def find_code_details_in_pending_saves(code):
    details = []
    for item in load_pending_saves():
        codes = item.get("codes", [])
        if code not in codes:
            continue
        order = item.get("order", {})
        details.append({
            "row_number": order.get("_row_number") or "локальная очередь",
            "date": order.get("Дата получения заказа", ""),
            "payment": order.get("Тип оплаты", ""),
            "client": order.get("Клиент", ""),
            "address": order.get("Адрес", ""),
            "representative": order.get("Торговый представитель", ""),
            "product": order.get("Товары", ""),
            "quantity": order.get("Кол-во ШТ", ""),
            "blocks": order.get("Кол-во блок", ""),
            "status": order.get(STATUS_COLUMN, "ожидает записи"),
            "codes_count": len(codes),
        })
    return details

def format_duplicate_code_details(code, details, current_order=None):
    lines = [
        "Дублирующийся КИЗ",
        f"Код: {code}",
    ]

    if current_order:
        lines.extend([
            "",
            "Текущая попытка:",
            f"Клиент: {current_order.get('Клиент', '')}",
            f"Тип оплаты: {current_order.get('Тип оплаты', '')}",
            f"Адрес: {current_order.get('Адрес', '')}",
            f"Товар: {current_order.get('Товары', '')}",
            f"Торговый представитель: {current_order.get('Торговый представитель', '')}",
        ])

    if not details:
        lines.extend([
            "",
            "Где найден: код есть в кэше уже принятых КИЗов, но строку Google Sheets определить не удалось.",
        ])
        return "\n".join(lines)

    lines.extend(["", "Где уже занят:"])
    for detail in details[:10]:
        lines.extend([
            f"Строка Google Sheets: {detail.get('row_number')}",
            f"Дата: {detail.get('date')}",
            f"Клиент: {detail.get('client')}",
            f"Тип оплаты: {detail.get('payment')}",
            f"Адрес: {detail.get('address')}",
            f"Товар: {detail.get('product')}",
            f"Торговый представитель: {detail.get('representative')}",
            f"Кол-во ШТ: {detail.get('quantity')}",
            f"План блоков: {detail.get('blocks')}",
            f"Статус: {detail.get('status')}",
            f"Кодов в строке: {detail.get('codes_count')}",
            "",
        ])
    if len(details) > 10:
        lines.append(f"Еще совпадений: {len(details) - 10}")
    return "\n".join(lines).strip()

def get_today_orders():
    try:
        client = get_google_client()
        sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)
        ensure_import_sheet_layout(sheet)
        all_rows = sheet.get_all_values()
        if not all_rows:
            raise ValueError("Лист Google Sheets пустой")

        header = [normalize_header_name(col) for col in all_rows[0]]
        header_idx, missing = validate_sheet_header(header)
        if missing:
            raise ValueError("В таблице не найдены обязательные колонки: " + ", ".join(missing))

        today_orders = []
        status_idx = header_idx.get(STATUS_COLUMN)
        status_updates = []

        for row_number, row in enumerate(all_rows[1:], start=2):
            if not any(normalize_text(cell) for cell in row):
                continue

            record = {}
            for col_name, idx in header_idx.items():
                record[col_name] = get_cell(row, idx)

            normalized_date = parse_date_to_standard(record.get("Дата получения заказа"))
            scanned_codes = split_codes(record.get("Отсканированные коды"))
            current_status = normalize_text(record.get(STATUS_COLUMN))
            calculated_status = get_order_status(record)
            if status_idx is not None and (
                not current_status
                or (calculated_status == STATUS_COMPLETED and not is_completed_status(current_status))
            ):
                status_updates.append({
                    "range": f"{column_index_to_letter(status_idx)}{row_number}",
                    "values": [[calculated_status]],
                })
                record[STATUS_COLUMN] = calculated_status

            if is_order_active(record):
                record["_row_number"] = row_number
                record["_normalized_date"] = normalized_date
                record["_existing_scanned_codes"] = scanned_codes
                today_orders.append(record)

        if status_updates:
            sheet.batch_update(status_updates, value_input_option="USER_ENTERED")

        return today_orders, sheet

    except Exception as e:
        logging.exception("Не удалось загрузить данные из Google Sheets")
        raise

def fetch_sheet_data():
    today_orders, sheet = get_today_orders()
    all_existing_codes = get_all_existing_codes(sheet) if sheet else set()
    all_existing_codes.update(get_pending_codes())
    return today_orders, sheet, all_existing_codes

def update_scanned_codes_to_gsheet(sheet, order, scanned_codes):
    try:
        if not scanned_codes:
            return False, "Нет отсканированных кодов для записи"

        if len(scanned_codes) != len(set(scanned_codes)):
            return False, "В текущей позиции есть повторяющиеся коды"

        ensure_import_sheet_layout(sheet)
        all_rows = sheet.get_all_values()
        if not all_rows:
            return False, "Лист Google Sheets пустой"

        header_idx, missing = validate_sheet_header(all_rows[0])
        if missing:
            return False, "В таблице не найдены обязательные колонки: " + ", ".join(missing)

        codes_idx = header_idx["Отсканированные коды"]
        target_row_number = parse_int_value(order.get("_row_number"))

        target_row = None
        if 2 <= target_row_number <= len(all_rows):
            candidate = all_rows[target_row_number - 1]
            if row_matches_order(candidate, header_idx, order):
                target_row = target_row_number

        if target_row is None:
            for row_number, row in enumerate(all_rows[1:], start=2):
                if row_matches_order(row, header_idx, order):
                    target_row = row_number
                    break

        if target_row is None:
            return False, "Не найдена строка заказа для записи кодов"

        existing_codes = split_codes(get_cell(all_rows[target_row - 1], codes_idx))
        if existing_codes:
            existing_set = set(existing_codes)
            scanned_set = set(scanned_codes)
            if not existing_set.issubset(scanned_set):
                return False, "В строке заказа уже есть другие отсканированные коды"

        duplicate_codes = []
        scanned_set = set(scanned_codes)
        for row_number, row in enumerate(all_rows[1:], start=2):
            if row_number == target_row:
                continue
            row_codes = set(split_codes(get_cell(row, codes_idx)))
            duplicates = scanned_set.intersection(row_codes)
            duplicate_codes.extend(sorted(duplicates))

        if duplicate_codes:
            return False, "Коды уже есть в другой строке Google Sheets: " + ", ".join(duplicate_codes[:3])

        status_idx = header_idx.get(STATUS_COLUMN)
        updated_order = dict(order)
        updated_order["Отсканированные коды"] = "\n".join(scanned_codes)
        updates = [{
            "range": f"{column_index_to_letter(codes_idx)}{target_row}",
            "values": [["\n".join(scanned_codes)]],
        }]
        if status_idx is not None:
            updates.append({
                "range": f"{column_index_to_letter(status_idx)}{target_row}",
                "values": [[get_order_status(updated_order)]],
            })
        sheet.batch_update(updates, value_input_option="USER_ENTERED")
        return True, "Коды записаны в Google Sheets"
    except Exception as e:
        logging.exception("Не удалось записать коды в Google Sheets")
        return False, str(e)

def build_day_report_rows_from_gsheet(sheet):
    all_rows = sheet.get_all_values()
    if not all_rows:
        return {"terminal": [], "transfer": [], "unknown": []}

    header_idx, missing = validate_sheet_header(all_rows[0])
    if missing:
        raise ValueError("В таблице не найдены обязательные колонки: " + ", ".join(missing))

    today_str = datetime.now().strftime("%d.%m.%Y")
    report_rows = {"terminal": [], "transfer": [], "unknown": []}

    for row in all_rows[1:]:
        if parse_date_to_standard(get_cell(row, header_idx.get("Дата получения заказа"))) != today_str:
            continue

        codes = split_codes(get_cell(row, header_idx.get("Отсканированные коды")))
        if not codes:
            continue

        payment_type = get_cell(row, header_idx.get("Тип оплаты"))
        payment_group = normalize_payment_type(payment_type)
        rows = report_rows[payment_group]

        for code in codes:
            pieces_per_block = get_product_rule(get_cell(row, header_idx.get("Товары")))["pieces_per_block"]
            rows.append({
                "Клиент": get_cell(row, header_idx.get("Клиент")),
                "Торговый представитель": get_cell(row, header_idx.get("Торговый представитель")),
                "Адрес": get_cell(row, header_idx.get("Адрес")),
                "Товар": get_cell(row, header_idx.get("Товары")),
                "Тип оплаты": payment_type,
                "Кол-во ШТ в блоке": pieces_per_block,
                "Кол-во блок": 1,
                "Итого ШТ": pieces_per_block,
                "Код": code
            })

    return report_rows

def add_pending_saves_to_report_rows(report_rows):
    today_str = datetime.now().strftime("%d.%m.%Y")
    existing_codes = {
        row.get("Код")
        for rows in report_rows.values()
        for row in rows
        if row.get("Код")
    }

    for item in load_pending_saves():
        order = item.get("order", {})
        if parse_date_to_standard(order.get("Дата получения заказа")) != today_str:
            continue

        payment_type = order.get("Тип оплаты", "")
        payment_group = normalize_payment_type(payment_type)
        rows = report_rows[payment_group]
        pieces_per_block = get_product_rule(order.get("Товары"))["pieces_per_block"]

        for code in item.get("codes", []):
            if not code or code in existing_codes:
                continue
            rows.append({
                "Клиент": order.get("Клиент", ""),
                "Торговый представитель": order.get("Торговый представитель", ""),
                "Адрес": order.get("Адрес", ""),
                "Товар": order.get("Товары", ""),
                "Тип оплаты": payment_type,
                "Кол-во ШТ в блоке": pieces_per_block,
                "Кол-во блок": 1,
                "Итого ШТ": pieces_per_block,
                "Код": code,
            })
            existing_codes.add(code)

    return report_rows

def parse_import_day(value):
    text = normalize_text(value)
    if not text:
        return ""
    return parse_date_to_standard(text.split()[0]) or text

def parse_datetime_for_sort(value):
    text = normalize_text(value)
    for fmt in ("%d.%m.%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return datetime.min

def split_source_files(value):
    text = normalize_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part.strip()]

def document_report_key(source_file, import_day):
    return make_hash({
        "source_file": normalize_text(source_file),
        "import_day": normalize_text(import_day),
    })

def pending_codes_for_order(order, pending_saves=None):
    pending_saves = pending_saves if pending_saves is not None else load_pending_saves()
    order_id = normalize_text(order.get("ID заказа"))
    row_number = normalize_text(order.get("_row_number"))
    codes = []
    seen = set()

    for item in pending_saves:
        pending_order = item.get("order", {})
        pending_order_id = normalize_text(pending_order.get("ID заказа"))
        pending_row_number = normalize_text(pending_order.get("_row_number"))
        matches = False
        if order_id and pending_order_id and order_id == pending_order_id:
            matches = True
        elif row_number and pending_row_number and row_number == pending_row_number:
            matches = True

        if not matches:
            continue

        for code in item.get("codes", []):
            code = normalize_text(code)
            if code and code not in seen:
                codes.append(code)
                seen.add(code)

    return codes

def merge_order_codes_with_pending(order, sheet_codes, pending_saves=None):
    codes = list(sheet_codes)
    seen = set(codes)
    for code in pending_codes_for_order(order, pending_saves=pending_saves):
        if code not in seen:
            codes.append(code)
            seen.add(code)
    return codes

def iter_document_orders_from_rows(all_rows, pending_saves=None):
    if not all_rows:
        return []

    header_idx, missing = validate_sheet_header(all_rows[0])
    if missing:
        raise ValueError("В таблице не найдены обязательные колонки: " + ", ".join(missing))

    pending_saves = pending_saves if pending_saves is not None else load_pending_saves()
    rows = []
    for row_number, row in enumerate(all_rows[1:], start=2):
        source_files = split_source_files(get_cell(row, header_idx.get("Источник файла")))
        if not source_files:
            continue

        order = {col_name: get_cell(row, idx) for col_name, idx in header_idx.items()}
        order["_row_number"] = row_number
        sheet_codes = split_codes(order.get("Отсканированные коды"))
        codes = merge_order_codes_with_pending(order, sheet_codes, pending_saves=pending_saves)
        plan_blocks = get_plan_blocks(order)
        import_at = order.get("Дата импорта", "")
        rows.append({
            "order": order,
            "row_number": row_number,
            "source_files": source_files,
            "import_at": import_at,
            "import_day": parse_import_day(import_at),
            "plan_blocks": plan_blocks,
            "codes": codes,
            "sheet_codes_count": len(sheet_codes),
            "pending_codes_count": max(0, len(codes) - len(sheet_codes)),
        })
    return rows

def build_document_summaries_from_gsheet(sheet, limit=12):
    document_rows = iter_document_orders_from_rows(sheet.get_all_values())
    documents = {}
    for item in document_rows:
        for source_file in item["source_files"]:
            key = document_report_key(source_file, item["import_day"])
            document = documents.setdefault(key, {
                "key": key,
                "source_file": source_file,
                "import_day": item["import_day"],
                "last_import": item["import_at"],
                "positions": 0,
                "completed_positions": 0,
                "plan_blocks": 0,
                "scanned_blocks": 0,
                "pending_blocks": 0,
            })
            document["positions"] += 1
            document["plan_blocks"] += item["plan_blocks"]
            document["scanned_blocks"] += len(item["codes"])
            document["pending_blocks"] += item["pending_codes_count"]
            if item["plan_blocks"] > 0 and len(item["codes"]) >= item["plan_blocks"]:
                document["completed_positions"] += 1
            if parse_datetime_for_sort(item["import_at"]) > parse_datetime_for_sort(document["last_import"]):
                document["last_import"] = item["import_at"]

    summaries = sorted(
        documents.values(),
        key=lambda document: parse_datetime_for_sort(document.get("last_import")),
        reverse=True,
    )
    return summaries[:limit] if limit else summaries

def truncate_middle(text, max_length):
    text = normalize_text(text)
    if len(text) <= max_length:
        return text
    if max_length <= 3:
        return text[:max_length]
    head = max_length // 2
    tail = max_length - head - 3
    return text[:head] + "..." + text[-tail:]

def create_document_report_excel(sheet, document_key):
    import pandas as pd

    document_rows = iter_document_orders_from_rows(sheet.get_all_values())
    selected = []
    selected_source = ""
    selected_import_day = ""
    for item in document_rows:
        for source_file in item["source_files"]:
            if document_report_key(source_file, item["import_day"]) != document_key:
                continue
            selected.append((source_file, item))
            selected_source = source_file
            selected_import_day = item["import_day"]

    if not selected:
        return {"empty": True, "document_key": document_key}

    positions = []
    codes_rows = []
    missing_rows = []
    total_plan = 0
    total_scanned = 0
    completed_positions = 0
    pending_count = 0
    last_import = ""

    for source_file, item in selected:
        order = item["order"]
        plan_blocks = item["plan_blocks"]
        codes = item["codes"]
        scanned_count = len(codes)
        remaining = max(0, plan_blocks - scanned_count)
        total_plan += plan_blocks
        total_scanned += scanned_count
        pending_count += item["pending_codes_count"]
        if plan_blocks > 0 and scanned_count >= plan_blocks:
            completed_positions += 1
        if parse_datetime_for_sort(item["import_at"]) > parse_datetime_for_sort(last_import):
            last_import = item["import_at"]

        position_row = {
            "Документ": source_file,
            "Дата импорта": item["import_at"],
            "Строка Google Sheets": item["row_number"],
            "Строка файла": order.get("Строка файла", ""),
            "Дата заказа": order.get("Дата получения заказа", ""),
            "Клиент": order.get("Клиент", ""),
            "Тип оплаты": order.get("Тип оплаты", ""),
            "Адрес": order.get("Адрес", ""),
            "Торговый представитель": order.get("Торговый представитель", ""),
            "Товар": order.get("Товары", ""),
            "План КИЗ": plan_blocks,
            "Отсканировано КИЗ": scanned_count,
            "Осталось КИЗ": remaining,
            "КИЗ в локальной очереди": item["pending_codes_count"],
            "Статус": "Выполнено" if plan_blocks > 0 and scanned_count >= plan_blocks else "Не выполнено",
        }
        positions.append(position_row)
        if remaining:
            missing_rows.append(position_row.copy())

        for code in codes:
            codes_rows.append({
                "Документ": source_file,
                "Дата импорта": item["import_at"],
                "Строка Google Sheets": item["row_number"],
                "Строка файла": order.get("Строка файла", ""),
                "Клиент": order.get("Клиент", ""),
                "Тип оплаты": order.get("Тип оплаты", ""),
                "Адрес": order.get("Адрес", ""),
                "Товар": order.get("Товары", ""),
                "КИЗ": code,
            })

    completion_percent = round((total_scanned / total_plan) * 100, 1) if total_plan else 0
    summary_rows = [{
        "Документ": selected_source,
        "Дата импорта": last_import or selected_import_day,
        "Позиций": len(positions),
        "Позиций выполнено": completed_positions,
        "План КИЗ": total_plan,
        "Отсканировано КИЗ": total_scanned,
        "Осталось КИЗ": max(0, total_plan - total_scanned),
        "КИЗ в локальной очереди": pending_count,
        "Готовность, %": completion_percent,
    }]

    os.makedirs(REPORTS_DIR, exist_ok=True)
    safe_name = re.sub(r"[^A-Za-zА-Яа-я0-9_.-]+", "_", selected_source)[:40] or "document"
    filename = os.path.join(
        REPORTS_DIR,
        f"document_report_{safe_name}_{datetime.now().strftime('%d.%m.%Y_%H%M%S')}.xlsx",
    )

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Сводка", index=False)
        pd.DataFrame(positions).to_excel(writer, sheet_name="Позиции", index=False)
        if codes_rows:
            pd.DataFrame(codes_rows).to_excel(writer, sheet_name="КИЗы", index=False)
        else:
            pd.DataFrame({"Сообщение": ["По документу пока нет отсканированных КИЗов"]}).to_excel(
                writer,
                sheet_name="КИЗы",
                index=False,
            )
        if missing_rows:
            pd.DataFrame(missing_rows).to_excel(writer, sheet_name="Недосканировано", index=False)
        else:
            pd.DataFrame({"Сообщение": ["Все позиции документа выполнены"]}).to_excel(
                writer,
                sheet_name="Недосканировано",
                index=False,
            )

    return {
        "empty": False,
        "filename": filename,
        "source_file": selected_source,
        "import_day": selected_import_day,
        "last_import": last_import,
        "positions": len(positions),
        "completed_positions": completed_positions,
        "plan_blocks": total_plan,
        "scanned_blocks": total_scanned,
        "remaining_blocks": max(0, total_plan - total_scanned),
        "pending_blocks": pending_count,
        "completion_percent": completion_percent,
    }

def create_day_report_excel(sheet, filename=None, include_pending=True):
    import pandas as pd

    report_rows = build_day_report_rows_from_gsheet(sheet)
    if include_pending:
        report_rows = add_pending_saves_to_report_rows(report_rows)
    terminal_rows = report_rows["terminal"]
    transfer_rows = report_rows["transfer"]
    unknown_rows = report_rows["unknown"]
    total_report_rows = len(terminal_rows) + len(transfer_rows) + len(unknown_rows)

    result = {
        "empty": total_report_rows == 0,
        "filename": filename,
        "total_report_rows": total_report_rows,
        "terminal_count": len(terminal_rows),
        "transfer_count": len(transfer_rows),
        "unknown_count": len(unknown_rows),
    }
    if total_report_rows == 0:
        return result

    os.makedirs(REPORTS_DIR, exist_ok=True)
    if not filename:
        filename = os.path.join(REPORTS_DIR, f"scan_report_{datetime.now().strftime('%d.%m.%Y_%H%M%S')}.xlsx")
        result["filename"] = filename

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:
        if terminal_rows:
            pd.DataFrame(terminal_rows).to_excel(writer, sheet_name="Терминал", index=False)
        else:
            pd.DataFrame({"Сообщение": ["Нет данных"]}).to_excel(writer, sheet_name="Терминал", index=False)

        if transfer_rows:
            pd.DataFrame(transfer_rows).to_excel(writer, sheet_name="Перечисление", index=False)
        else:
            pd.DataFrame({"Сообщение": ["Нет данных"]}).to_excel(writer, sheet_name="Перечисление", index=False)

        if unknown_rows:
            pd.DataFrame(unknown_rows).to_excel(writer, sheet_name="Не распознано", index=False)

    return result

def build_summary_products_from_gsheet(sheet, group_key):
    all_rows = sheet.get_all_values()
    if not all_rows:
        return []

    header_idx, missing = validate_sheet_header(all_rows[0])
    if missing:
        raise ValueError("В таблице не найдены обязательные колонки: " + ", ".join(missing))

    products = []
    for row in all_rows[1:]:
        row_group = (
            get_cell(row, header_idx.get("Клиент")) or "Клиент не указан",
            get_cell(row, header_idx.get("Тип оплаты")) or "Оплата не указана",
            get_cell(row, header_idx.get("Адрес")) or "Адрес не указан",
        )
        if row_group != group_key:
            continue

        codes = split_codes(get_cell(row, header_idx.get("Отсканированные коды")))
        if not codes:
            continue

        products.append({
            "Клиент": get_cell(row, header_idx.get("Клиент")),
            "Адрес": get_cell(row, header_idx.get("Адрес")),
            "Торговый представитель": get_cell(row, header_idx.get("Торговый представитель")),
            "Товары": get_cell(row, header_idx.get("Товары")),
            "Тип оплаты": get_cell(row, header_idx.get("Тип оплаты")),
            "Кол-во ШТ в блоке": get_product_rule(get_cell(row, header_idx.get("Товары")))["pieces_per_block"],
            "План": parse_int_value(get_cell(row, header_idx.get("Кол-во блок"))),
            "Отсканировано": len(codes),
            "Коды": codes,
        })

    return products

def load_font(candidates, size):
    for path in candidates:
        try:
            return ImageFont.truetype(path, size=size)
        except:
            continue
    return ImageFont.load_default()

def wrap_text(text, max_chars):
    if len(text) <= max_chars:
        return [text]
    words = text.split()
    lines = []
    current_line = ""
    for word in words:
        if len(current_line) + len(word) + 1 <= max_chars:
            if current_line:
                current_line += " " + word
            else:
                current_line = word
        else:
            if current_line:
                lines.append(current_line)
            current_line = word
    if current_line:
        lines.append(current_line)
    return lines

def mm_to_px(mm_value):
    return int(mm_value / 25.4 * LABEL_DPI)

def powershell_quote(value):
    return "'" + str(value).replace("'", "''") + "'"

def send_image_to_windows_printer(file_path, printer_name=""):
    image_path = os.path.abspath(file_path)
    printer_name = normalize_text(printer_name)
    paper_width = int(round(LABEL_WIDTH_MM / 25.4 * 100))
    paper_height = int(round(LABEL_HEIGHT_MM / 25.4 * 100))
    printer_line = ""
    if printer_name and printer_name != "Термопринтер":
        printer_line = f"$printDocument.PrinterSettings.PrinterName = {powershell_quote(printer_name)}"

    script = f"""
Add-Type -AssemblyName System.Drawing
$imagePath = {powershell_quote(image_path)}
$image = [System.Drawing.Image]::FromFile($imagePath)
$printDocument = New-Object System.Drawing.Printing.PrintDocument
{printer_line}
$printDocument.DocumentName = "{APP_NAME} summary"
$printDocument.DefaultPageSettings.PaperSize = New-Object System.Drawing.Printing.PaperSize("Label100x100", {paper_width}, {paper_height})
$printDocument.DefaultPageSettings.Margins = New-Object System.Drawing.Printing.Margins(0, 0, 0, 0)
$printDocument.OriginAtMargins = $false
$printDocument.add_PrintPage({{
    param($sender, $event)
    $event.Graphics.DrawImage($image, $event.PageBounds)
    $event.HasMorePages = $false
}})
try {{
    $printDocument.Print()
}} finally {{
    $image.Dispose()
    $printDocument.Dispose()
}}
"""
    ps_file = tempfile.NamedTemporaryFile(suffix=".ps1", delete=False, mode="w", encoding="utf-8")
    try:
        ps_file.write(script)
        ps_file.close()
        creationflags = subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0
        subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                ps_file.name,
            ],
            check=True,
            timeout=30,
            creationflags=creationflags,
        )
        return True
    finally:
        try:
            os.remove(ps_file.name)
        except OSError:
            pass

def send_image_to_printer(file_path, printer_name=""):
    try:
        if os.name == 'nt':
            return send_image_to_windows_printer(file_path, printer_name=printer_name)

        command = ["lp", "-o", f"media=Custom.{LABEL_WIDTH_MM}x{LABEL_HEIGHT_MM}mm", file_path]
        if normalize_text(printer_name) and printer_name != "Термопринтер":
            command[1:1] = ["-d", printer_name]
        subprocess.run(command, check=True, timeout=30)
        return True
    except Exception:
        logging.exception("Не удалось отправить сводку напрямую на печать")
        return False

def print_summary(address, all_products):
    try:
        if not all_products:
            return None

        width_px = mm_to_px(LABEL_WIDTH_MM)
        height_px = mm_to_px(LABEL_HEIGHT_MM)
        scale = LABEL_DPI / 96

        def s(value):
            return int(value * scale)

        products_per_page = 3
        pages = []
        for i in range(0, len(all_products), products_per_page):
            pages.append(all_products[i:i + products_per_page])

        print_settings = load_print_settings()
        printer_name = normalize_text(print_settings.get("printer_name"))
        printed_files = []

        for page_idx, page_products in enumerate(pages):
            img = Image.new("RGB", (width_px, height_px), "white")
            draw = ImageDraw.Draw(img)

            margin = s(10)
            inner_width = width_px - (margin * 2)
            inner_height = height_px - (margin * 2)

            draw.rectangle([margin, margin, width_px - margin, height_px - margin], outline="#333333", width=max(1, s(2)))

            font_title = load_font(["arialbd.ttf", "Arial Bold.ttf"], s(14))
            font_text = load_font(["arial.ttf", "Arial.ttf"], s(11))
            font_small = load_font(["arial.ttf", "Arial.ttf"], s(9))
            font_bold = load_font(["arialbd.ttf", "Arial Bold.ttf"], s(12))

            y = margin + s(8)
            x = margin + s(8)
            line_height = s(18)

            draw.text((x, y), f"СВОДНЫЙ ОТЧЁТ ПО АДРЕСУ", fill="black", font=font_title)
            y += line_height + s(5)

            addr_lines = wrap_text(address, 34)
            for line in addr_lines:
                draw.text((x, y), line, fill="black", font=font_small)
                y += line_height - s(2)
            y += s(5)

            if all_products:
                client = all_products[0].get('Клиент', '')
                draw.text((x, y), f"Клиент: {client[:35]}", fill="black", font=font_text)
                y += line_height

            if all_products:
                rep = all_products[0].get('Торговый представитель', '')
                draw.text((x, y), f"Торг.пред: {rep[:35]}", fill="black", font=font_text)
                y += line_height + s(5)

            draw.line([(x, y), (x + inner_width - s(16), y)], fill="#cccccc", width=max(1, s(1)))
            y += s(10)

            page_text = f"Стр. {page_idx + 1} из {len(pages)}"
            draw.text((width_px - margin - s(60), margin + s(8)), page_text, fill="gray", font=font_small)

            draw.text((x, y), "№", fill="black", font=font_bold)
            draw.text((x + s(25), y), "Товар", fill="black", font=font_bold)
            draw.text((x + s(200), y), "Блоков", fill="black", font=font_bold)
            draw.text((x + s(260), y), "ШТ", fill="black", font=font_bold)
            y += line_height + s(3)
            draw.line([(x, y - s(5)), (x + inner_width - s(16), y - s(5))], fill="#000000", width=max(1, s(1)))

            total_blocks = 0
            total_shields = 0

            for idx, product in enumerate(page_products):
                product_name = product.get('Товары', '')[:22]
                blocks = product.get('Отсканировано', 0)
                pieces_per_block = parse_int_value(product.get("Кол-во ШТ в блоке")) or DEFAULT_PIECES_PER_BLOCK
                shields = blocks * pieces_per_block
                total_blocks += blocks
                total_shields += shields

                draw.text((x, y), f"{idx + 1 + (page_idx * products_per_page)}", fill="black", font=font_text)
                draw.text((x + s(25), y), product_name, fill="black", font=font_text)
                draw.text((x + s(205), y), str(blocks), fill="black", font=font_text)
                draw.text((x + s(265), y), str(shields), fill="black", font=font_text)
                y += line_height

                if y > height_px - s(60) and idx < len(page_products) - 1:
                    draw.text((x, height_px - margin - s(12)), datetime.now().strftime("%d.%m.%Y %H:%M"), fill="gray", font=font_small)
                    break

            draw.line([(x, y), (x + inner_width - s(16), y)], fill="#cccccc", width=max(1, s(1)))
            y += s(8)

            draw.text((x, y), f"ИТОГО:", fill="black", font=font_bold)
            draw.text((x + s(205), y), str(total_blocks), fill="black", font=font_bold)
            draw.text((x + s(265), y), str(total_shields), fill="black", font=font_bold)

            draw.text((x, height_px - margin - s(12)), datetime.now().strftime("%d.%m.%Y %H:%M"), fill="gray", font=font_small)

            temp_file = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
            img.save(temp_file.name, dpi=(LABEL_DPI, LABEL_DPI))
            temp_file.close()
            printed_files.append(temp_file.name)

            if not send_image_to_printer(temp_file.name, printer_name=printer_name):
                return None

        return printed_files
    except Exception as e:
        logging.exception("Ошибка печати сводного листа")
        return None

def write_scan_backup(action, order, code=None, codes=None):
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        filename = os.path.join(BACKUP_DIR, f"scan_backup_{datetime.now().strftime('%d.%m.%Y')}.jsonl")
        payload = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "action": action,
            "row_number": order.get("_row_number"),
            "date": order.get("Дата получения заказа", ""),
            "client": order.get("Клиент", ""),
            "address": order.get("Адрес", ""),
            "product": order.get("Товары", ""),
            "payment_type": order.get("Тип оплаты", ""),
            "code": code,
            "codes": codes or [],
        }
        with open(filename, "a", encoding="utf-8") as backup_file:
            backup_file.write(json.dumps(payload, ensure_ascii=False) + "\n")
        return True
    except Exception:
        logging.exception("Не удалось записать локальный backup")
        return False

def load_pending_prints():
    try:
        if not os.path.exists(PENDING_PRINTS_FILE):
            return []
        with open(PENDING_PRINTS_FILE, "r", encoding="utf-8") as pending_file:
            data = json.load(pending_file)
        if isinstance(data, list):
            return data
        return []
    except Exception:
        logging.exception("Не удалось загрузить очередь печати")
        return []

def save_pending_prints(items):
    try:
        with open(PENDING_PRINTS_FILE, "w", encoding="utf-8") as pending_file:
            json.dump(items, pending_file, ensure_ascii=False, indent=2)
        return True
    except Exception:
        logging.exception("Не удалось сохранить очередь печати")
        return False

def make_pending_print_id(address, products):
    payload = {
        "address": address,
        "products": [
            {
                "client": product.get("Клиент", ""),
                "address": product.get("Адрес", ""),
                "product": product.get("Товары", ""),
                "codes": product.get("Коды", []),
            }
            for product in products
        ],
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()

def add_pending_print(address, products):
    pending = load_pending_prints()
    pending_id = make_pending_print_id(address, products)
    for item in pending:
        if item.get("id") == pending_id:
            return pending_id

    pending.append({
        "id": pending_id,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "address": address,
        "products": products,
    })
    save_pending_prints(pending)
    return pending_id

def remove_pending_print(pending_id):
    pending = load_pending_prints()
    new_pending = [item for item in pending if item.get("id") != pending_id]
    if len(new_pending) != len(pending):
        save_pending_prints(new_pending)

def load_pending_saves():
    data = load_json_file(PENDING_SAVES_FILE, [])
    return data if isinstance(data, list) else []

def save_pending_saves(items):
    return save_json_file(PENDING_SAVES_FILE, items)

def make_pending_save_id(order, scanned_codes):
    return make_hash({
        "order_id": order.get("ID заказа", ""),
        "row_number": order.get("_row_number", ""),
        "date": order.get("Дата получения заказа", ""),
        "client": order.get("Клиент", ""),
        "address": order.get("Адрес", ""),
        "product": order.get("Товары", ""),
        "codes": scanned_codes,
    })

def add_pending_save(order, scanned_codes, reason):
    pending = load_pending_saves()
    pending_id = make_pending_save_id(order, scanned_codes)
    for item in pending:
        if item.get("id") == pending_id:
            item["last_error"] = reason
            item["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            save_pending_saves(pending)
            return pending_id

    pending.append({
        "id": pending_id,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "order": {key: value for key, value in order.items() if not key.startswith("_existing")},
        "codes": scanned_codes,
        "last_error": reason,
    })
    save_pending_saves(pending)
    return pending_id

def remove_pending_save(pending_id):
    pending = load_pending_saves()
    new_pending = [item for item in pending if item.get("id") != pending_id]
    if len(new_pending) != len(pending):
        save_pending_saves(new_pending)

def get_pending_codes():
    codes = set()
    for item in load_pending_saves():
        for code in item.get("codes", []):
            if code:
                codes.add(code)
    return codes

def is_retryable_save_error(message):
    text = normalize_text(message).lower()
    non_retryable = [
        "повтор",
        "другой строке",
        "не найдена строка",
        "обязательные колонки",
        "уже есть другие",
        "лист google sheets пустой",
        "нет отсканированных кодов",
    ]
    if any(marker in text for marker in non_retryable):
        return False

    retryable = [
        "timeout",
        "timed out",
        "connection",
        "network",
        "temporary",
        "ssl",
        "socket",
        "503",
        "502",
        "500",
        "429",
        "quota",
        "unavailable",
        "service",
        "broken pipe",
    ]
    return any(marker in text for marker in retryable) or bool(text)

def sync_pending_saves(sheet=None):
    pending = load_pending_saves()
    if not pending:
        return {"synced": 0, "failed": 0, "remaining": 0}

    if sheet is None:
        client = get_google_client()
        sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)

    synced = 0
    failed = 0
    remaining = []
    for item in pending:
        order = item.get("order", {})
        codes = item.get("codes", [])
        ok, message = update_scanned_codes_to_gsheet(sheet, order, codes)
        if ok:
            synced += 1
            write_scan_backup("pending_save_synced", order, codes=codes)
            continue

        failed += 1
        item["last_error"] = message
        item["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        remaining.append(item)

    save_pending_saves(remaining)
    return {"synced": synced, "failed": failed, "remaining": len(remaining)}

def fetch_sheet_data_with_sync():
    today_orders, sheet = get_today_orders()
    sync_result = sync_pending_saves(sheet)
    if sync_result.get("synced"):
        today_orders, sheet = get_today_orders()
    all_existing_codes = get_all_existing_codes(sheet) if sheet else set()
    all_existing_codes.update(get_pending_codes())
    return today_orders, sheet, all_existing_codes, sync_result

def build_control_panel_stats_from_gsheet(sheet):
    all_rows = sheet.get_all_values()
    if not all_rows:
        return {}

    header_idx, missing = validate_sheet_header(all_rows[0])
    if missing:
        raise ValueError("В таблице не найдены обязательные колонки: " + ", ".join(missing))

    today_str = datetime.now().strftime("%d.%m.%Y")
    groups = {}
    products = {}
    payments = {"terminal": 0, "transfer": 0, "unknown": 0}
    positions = 0
    completed_positions = 0
    in_progress_positions = 0
    new_positions = 0
    plan_blocks = 0
    scanned_blocks = 0

    for row in all_rows[1:]:
        if parse_date_to_standard(get_cell(row, header_idx.get("Дата получения заказа"))) != today_str:
            continue

        positions += 1
        order = {col_name: get_cell(row, idx) for col_name, idx in header_idx.items()}
        group_key = order_group_key(order)
        groups.setdefault(group_key, {"positions": 0, "completed": 0})
        groups[group_key]["positions"] += 1

        blocks = get_plan_blocks(order)
        codes_count = len(split_codes(order.get("Отсканированные коды")))
        plan_blocks += blocks
        scanned_blocks += codes_count
        products[order.get("Товары", "Товар не указан")] = products.get(order.get("Товары", "Товар не указан"), 0) + blocks
        payments[normalize_payment_type(order.get("Тип оплаты"))] += 1

        if blocks > 0 and codes_count >= blocks:
            completed_positions += 1
            groups[group_key]["completed"] += 1
        elif codes_count > 0:
            in_progress_positions += 1
        else:
            new_positions += 1

    completed_groups = sum(1 for group in groups.values() if group["positions"] == group["completed"])
    active_groups = max(0, len(groups) - completed_groups)
    return {
        "positions": positions,
        "groups": len(groups),
        "active_groups": active_groups,
        "completed_groups": completed_groups,
        "completed_positions": completed_positions,
        "in_progress_positions": in_progress_positions,
        "new_positions": new_positions,
        "plan_blocks": plan_blocks,
        "scanned_blocks": scanned_blocks,
        "remaining_blocks": max(0, plan_blocks - scanned_blocks),
        "payments": payments,
        "products": dict(sorted(products.items(), key=lambda item: item[0].lower())),
        "pending_saves": len(load_pending_saves()),
        "pending_prints": len(load_pending_prints()),
    }

def load_print_settings():
    defaults = {
        "printer_name": "Термопринтер",
        "label_width_mm": LABEL_WIDTH_MM,
        "label_height_mm": LABEL_HEIGHT_MM,
        "dpi": LABEL_DPI,
        "scale": "100%",
    }
    settings = load_json_file(PRINT_SETTINGS_FILE, {})
    if isinstance(settings, dict):
        defaults.update({key: value for key, value in settings.items() if value not in (None, "")})
    return defaults

def save_print_settings(settings):
    return save_json_file(PRINT_SETTINGS_FILE, settings)

def load_telegram_settings():
    defaults = {
        "enabled": False,
        "bot_token": "",
        "chat_id": "",
        "chat_ids": [],
        "send_reports": True,
        "send_scan_backups": False,
        "send_pending_files": False,
        "send_error_log": True,
    }
    settings = load_json_file(TELEGRAM_SETTINGS_FILE, {})
    if isinstance(settings, dict):
        defaults.update({key: value for key, value in settings.items() if value is not None})
    return defaults

def get_telegram_chat_ids(settings=None):
    settings = settings or load_telegram_settings()
    chat_ids = []
    raw_chat_ids = settings.get("chat_ids", [])
    if isinstance(raw_chat_ids, list):
        chat_ids.extend(normalize_text(chat_id) for chat_id in raw_chat_ids)
    else:
        chat_ids.extend(
            normalize_text(chat_id)
            for chat_id in str(raw_chat_ids).split(",")
        )

    legacy_chat_id = normalize_text(settings.get("chat_id"))
    if legacy_chat_id:
        chat_ids.append(legacy_chat_id)

    unique_chat_ids = []
    seen = set()
    for chat_id in chat_ids:
        if chat_id and chat_id not in seen:
            unique_chat_ids.append(chat_id)
            seen.add(chat_id)
    return unique_chat_ids

def telegram_is_configured(settings=None):
    settings = settings or load_telegram_settings()
    return bool(settings.get("enabled") and normalize_text(settings.get("bot_token")) and get_telegram_chat_ids(settings))

def safe_telegram_document_path(path):
    if not path:
        return False
    normalized_path = os.path.abspath(path)
    blocked = {
        os.path.abspath(CREDENTIALS_FILE),
        os.path.abspath(TELEGRAM_SETTINGS_FILE),
        os.path.abspath(YANDEX_GEOCODER_KEY_FILE),
    }
    return os.path.exists(normalized_path) and normalized_path not in blocked

TELEGRAM_CALLBACK_TODAY_SCANS = "today_scans"
TELEGRAM_CALLBACK_TODAY_LOG = "today_log"
TELEGRAM_CALLBACK_DOCUMENTS = "documents"
TELEGRAM_CALLBACK_DOCUMENT_PREFIX = "doc:"

def load_telegram_state():
    state = load_json_file(TELEGRAM_STATE_FILE, {})
    return state if isinstance(state, dict) else {}

def save_telegram_state(state):
    return save_json_file(TELEGRAM_STATE_FILE, state)

def telegram_reports_keyboard():
    return {
        "inline_keyboard": [
            [{"text": "Скачать сканы за сегодня", "callback_data": TELEGRAM_CALLBACK_TODAY_SCANS}],
            [{"text": "Документы по импорту", "callback_data": TELEGRAM_CALLBACK_DOCUMENTS}],
            [{"text": "Скачать сегодняшний лог", "callback_data": TELEGRAM_CALLBACK_TODAY_LOG}],
        ]
    }

def telegram_documents_keyboard(document_summaries):
    keyboard = []
    for document in document_summaries:
        plan = document.get("plan_blocks", 0)
        scanned = document.get("scanned_blocks", 0)
        source_file = truncate_middle(document.get("source_file", "Документ"), 32)
        button_text = f"{source_file} | {scanned}/{plan}"
        keyboard.append([{
            "text": button_text,
            "callback_data": TELEGRAM_CALLBACK_DOCUMENT_PREFIX + document["key"],
        }])
    keyboard.append([{"text": "Назад", "callback_data": "menu"}])
    return {"inline_keyboard": keyboard}

def telegram_api_request(token, method_name, payload=None, timeout=30):
    payload = payload or {}
    encoded_payload = urllib.parse.urlencode(payload).encode("utf-8")
    request = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/{method_name}",
        data=encoded_payload,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        result = json.load(response)
    if result.get("ok"):
        return result.get("result")
    raise RuntimeError(normalize_text(result.get("description")) or "Telegram вернул ошибку")

def send_telegram_message_to_chat(chat_id, text, token, reply_markup=None):
    fields = {
        "chat_id": normalize_text(chat_id),
        "text": normalize_text(text)[:4096],
    }
    if reply_markup:
        fields["reply_markup"] = json.dumps(reply_markup, ensure_ascii=False)
    telegram_api_request(token, "sendMessage", fields, timeout=30)
    return True, "Отправлено в Telegram"

def send_telegram_message(text, reply_markup=None):
    settings = load_telegram_settings()
    if not telegram_is_configured(settings):
        return False, "Telegram не настроен"

    token = normalize_text(settings.get("bot_token"))
    errors = []
    sent = 0
    for chat_id in get_telegram_chat_ids(settings):
        try:
            send_telegram_message_to_chat(chat_id, text, token, reply_markup=reply_markup)
            sent += 1
        except Exception as exc:
            logging.exception("Не удалось отправить сообщение в Telegram")
            errors.append(f"{chat_id}: {exc}")

    if errors:
        return False, "; ".join(errors)
    return True, f"Отправлено получателям: {sent}"

def answer_telegram_callback_query(token, callback_query_id, text=""):
    if not callback_query_id:
        return
    fields = {
        "callback_query_id": callback_query_id,
        "text": normalize_text(text)[:200],
    }
    telegram_api_request(token, "answerCallbackQuery", fields, timeout=15)

def fetch_telegram_updates(token, offset=None):
    fields = {
        "timeout": 0,
        "allowed_updates": json.dumps(["message", "callback_query"]),
    }
    if offset:
        fields["offset"] = offset
    result = telegram_api_request(token, "getUpdates", fields, timeout=20)
    return result if isinstance(result, list) else []

def telegram_chat_is_authorized(chat_id, settings=None):
    return normalize_text(chat_id) in set(get_telegram_chat_ids(settings))

def telegram_document_file_name(document):
    return clean_file_name(document.get("file_name"), "telegram_import")

def telegram_document_is_supported_excel(document):
    return is_supported_excel_file_name(telegram_document_file_name(document))

def get_telegram_file_info(token, file_id):
    file_id = normalize_text(file_id)
    if not file_id:
        raise ValueError("Telegram не передал file_id документа")
    result = telegram_api_request(token, "getFile", {"file_id": file_id}, timeout=30)
    if not isinstance(result, dict) or not normalize_text(result.get("file_path")):
        raise RuntimeError("Telegram не вернул путь к файлу")
    return result

def download_telegram_file(token, file_path, destination_path):
    quoted_path = urllib.parse.quote(normalize_text(file_path), safe="/")
    request = urllib.request.Request(
        f"https://api.telegram.org/file/bot{token}/{quoted_path}",
        headers={"User-Agent": f"{APP_NAME}/{APP_VERSION}"},
    )
    with urllib.request.urlopen(request, timeout=TELEGRAM_FILE_DOWNLOAD_TIMEOUT_SECONDS) as response:
        with open(destination_path, "wb") as output_file:
            while True:
                chunk = response.read(1024 * 1024)
                if not chunk:
                    break
                output_file.write(chunk)

def download_telegram_document_to_temp(token, document):
    file_name = telegram_document_file_name(document)
    if not telegram_document_is_supported_excel(document):
        raise ValueError("Поддерживаются только Excel-файлы .xlsx и .xlsm")

    suffix = os.path.splitext(file_name)[1].lower() or ".xlsx"
    temp_file = tempfile.NamedTemporaryFile(prefix=f"{APP_NAME}_telegram_import_", suffix=suffix, delete=False)
    temp_path = temp_file.name
    temp_file.close()

    try:
        file_info = get_telegram_file_info(token, document.get("file_id"))
        download_telegram_file(token, file_info.get("file_path"), temp_path)
        return temp_path, file_name
    except Exception:
        try:
            os.remove(temp_path)
        except OSError:
            pass
        raise

def create_today_log_file():
    os.makedirs(REPORTS_DIR, exist_ok=True)
    today_prefix = datetime.now().strftime("%Y-%m-%d")
    output_path = os.path.join(REPORTS_DIR, f"{APP_NAME}_log_{today_prefix}.txt")
    entry_start_re = re.compile(r"^\d{4}-\d{2}-\d{2} ")
    selected_lines = []

    if os.path.exists(LOG_FILE):
        with open(LOG_FILE, "r", encoding="utf-8", errors="replace") as log_file:
            current_entry = []
            current_matches_today = False
            for line in log_file:
                if entry_start_re.match(line):
                    if current_entry and current_matches_today:
                        selected_lines.extend(current_entry)
                    current_entry = [line]
                    current_matches_today = line.startswith(today_prefix)
                else:
                    current_entry.append(line)
            if current_entry and current_matches_today:
                selected_lines.extend(current_entry)

    if not selected_lines:
        selected_lines = [f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} [INFO] Ошибок за сегодня в логе нет\n"]

    with open(output_path, "w", encoding="utf-8") as output_file:
        output_file.writelines(selected_lines)
    return output_path

def telegram_multipart_body(fields, file_field, file_path):
    boundary = f"----{APP_NAME}Boundary" + hashlib.sha1(str(datetime.now().timestamp()).encode("utf-8")).hexdigest()
    chunks = []
    for key, value in fields.items():
        chunks.append(f"--{boundary}\r\n".encode("utf-8"))
        chunks.append(f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode("utf-8"))
        chunks.append(str(value).encode("utf-8"))
        chunks.append(b"\r\n")

    filename = os.path.basename(file_path)
    chunks.append(f"--{boundary}\r\n".encode("utf-8"))
    chunks.append(
        f'Content-Disposition: form-data; name="{file_field}"; filename="{filename}"\r\n'.encode("utf-8")
    )
    chunks.append(b"Content-Type: application/octet-stream\r\n\r\n")
    with open(file_path, "rb") as file_obj:
        chunks.append(file_obj.read())
    chunks.append(b"\r\n")
    chunks.append(f"--{boundary}--\r\n".encode("utf-8"))
    return boundary, b"".join(chunks)

def send_telegram_document_to_chat(file_path, chat_id, caption, token):
    fields = {
        "chat_id": normalize_text(chat_id),
        "caption": normalize_text(caption)[:1024],
    }
    boundary, body = telegram_multipart_body(fields, "document", file_path)
    request = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendDocument",
        data=body,
        headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "Content-Length": str(len(body)),
        },
        method="POST",
    )
    with urllib.request.urlopen(request, timeout=30) as response:
        result = json.load(response)
    if result.get("ok"):
        return True, "Отправлено в Telegram"
    return False, normalize_text(result.get("description")) or "Telegram вернул ошибку"

def send_telegram_document(file_path, caption=""):
    settings = load_telegram_settings()
    if not telegram_is_configured(settings):
        return False, "Telegram не настроен"
    if not safe_telegram_document_path(file_path):
        return False, "Файл не найден или запрещён к отправке"

    token = normalize_text(settings.get("bot_token"))
    errors = []
    sent = 0
    for chat_id in get_telegram_chat_ids(settings):
        try:
            ok, message = send_telegram_document_to_chat(file_path, chat_id, caption, token)
            if ok:
                sent += 1
            else:
                errors.append(f"{chat_id}: {message}")
        except Exception as exc:
            logging.exception("Не удалось отправить документ в Telegram")
            errors.append(f"{chat_id}: {exc}")

    if errors:
        return False, "; ".join(errors)
    return True, f"Отправлено получателям: {sent}"

def load_pending_telegram():
    data = load_json_file(PENDING_TELEGRAM_FILE, [])
    return data if isinstance(data, list) else []

def save_pending_telegram(items):
    return save_json_file(PENDING_TELEGRAM_FILE, items)

def make_pending_telegram_id(file_path, caption):
    payload = {
        "path": os.path.abspath(file_path),
        "caption": normalize_text(caption),
    }
    return make_hash(payload)

def add_pending_telegram(file_path, caption, reason):
    if not safe_telegram_document_path(file_path):
        return None
    pending = load_pending_telegram()
    pending_id = make_pending_telegram_id(file_path, caption)
    for item in pending:
        if item.get("id") == pending_id:
            item["last_error"] = reason
            item["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            item["attempts"] = parse_int_value(item.get("attempts")) + 1
            save_pending_telegram(pending)
            return pending_id

    pending.append({
        "id": pending_id,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "path": os.path.abspath(file_path),
        "caption": normalize_text(caption),
        "attempts": 1,
        "last_error": reason,
    })
    save_pending_telegram(pending)
    return pending_id

def send_or_queue_telegram_document(file_path, caption):
    settings = load_telegram_settings()
    if not telegram_is_configured(settings):
        return False, "Telegram не настроен"
    ok, message = send_telegram_document(file_path, caption)
    if not ok:
        add_pending_telegram(file_path, caption, message)
    return ok, message

def sync_pending_telegram():
    if not telegram_is_configured():
        return {"sent": 0, "failed": 0, "remaining": len(load_pending_telegram())}

    pending = load_pending_telegram()
    sent = 0
    failed = 0
    remaining = []
    for item in pending:
        file_path = item.get("path")
        caption = item.get("caption", "")
        if not safe_telegram_document_path(file_path):
            continue
        ok, message = send_telegram_document(file_path, caption)
        if ok:
            sent += 1
            continue
        failed += 1
        item["last_error"] = message
        item["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        item["attempts"] = parse_int_value(item.get("attempts")) + 1
        remaining.append(item)
    save_pending_telegram(remaining)
    return {"sent": sent, "failed": failed, "remaining": len(remaining)}

def today_scan_backup_path():
    return os.path.join(BACKUP_DIR, f"scan_backup_{datetime.now().strftime('%d.%m.%Y')}.jsonl")

def collect_operational_documents(
    include_report=None,
    include_error_log=False,
    include_scan_backup=False,
    include_pending_files=False,
):
    settings = load_telegram_settings()
    documents = []
    if include_report and settings.get("send_reports"):
        documents.append((include_report, f"{APP_NAME}: Excel-отчёт за день"))
    if include_scan_backup and settings.get("send_scan_backups"):
        backup_path = today_scan_backup_path()
        if safe_telegram_document_path(backup_path):
            documents.append((backup_path, f"{APP_NAME}: backup сканирования за день"))
    if include_pending_files and settings.get("send_pending_files"):
        for path, caption in (
            (PENDING_SAVES_FILE, f"{APP_NAME}: очередь записи в Google Sheets"),
            (PENDING_PRINTS_FILE, f"{APP_NAME}: очередь печати сводок"),
            (PENDING_TELEGRAM_FILE, f"{APP_NAME}: очередь отправки в Telegram"),
        ):
            if safe_telegram_document_path(path):
                documents.append((path, caption))
    if include_error_log and settings.get("send_error_log") and safe_telegram_document_path(LOG_FILE):
        documents.append((LOG_FILE, f"{APP_NAME}: журнал ошибок"))
    return documents

def parse_version_parts(version):
    parts = re.findall(r"\d+", normalize_text(version))
    if not parts:
        return (0,)
    return tuple(int(part) for part in parts[:4])

def compare_versions(left, right):
    left_parts = parse_version_parts(left)
    right_parts = parse_version_parts(right)
    max_len = max(len(left_parts), len(right_parts))
    left_parts = left_parts + (0,) * (max_len - len(left_parts))
    right_parts = right_parts + (0,) * (max_len - len(right_parts))
    if left_parts < right_parts:
        return -1
    if left_parts > right_parts:
        return 1
    return 0

def fetch_update_info():
    if not UPDATE_INFO_URL:
        return None

    separator = "&" if "?" in UPDATE_INFO_URL else "?"
    url = f"{UPDATE_INFO_URL}{separator}_={int(datetime.now().timestamp())}"
    request = urllib.request.Request(
        url,
        headers={
            "Accept": "application/json",
            "User-Agent": f"{APP_NAME}/{APP_VERSION}",
        },
    )
    with urllib.request.urlopen(request, timeout=UPDATE_CHECK_TIMEOUT_SECONDS) as response:
        update_info = json.load(response)

    if not isinstance(update_info, dict):
        raise ValueError("Файл обновления должен быть JSON-объектом")
    return update_info

def file_sha256(path):
    digest = hashlib.sha256()
    with open(path, "rb") as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()

def download_update_file(update_info):
    download_url = normalize_text(update_info.get("download_url"))
    if not download_url:
        raise ValueError("В version.json не указан download_url для нового exe")

    parsed_url = urllib.parse.urlparse(download_url)
    suffix = os.path.splitext(parsed_url.path)[1] or ".exe"
    temp_file = tempfile.NamedTemporaryFile(prefix=f"{APP_NAME}_update_", suffix=suffix, delete=False)
    temp_path = temp_file.name
    temp_file.close()

    request = urllib.request.Request(
        download_url,
        headers={"User-Agent": f"{APP_NAME}/{APP_VERSION}"},
    )
    try:
        with urllib.request.urlopen(request, timeout=UPDATE_DOWNLOAD_TIMEOUT_SECONDS) as response:
            with open(temp_path, "wb") as file_obj:
                while True:
                    chunk = response.read(1024 * 1024)
                    if not chunk:
                        break
                    file_obj.write(chunk)

        expected_sha256 = normalize_text(update_info.get("sha256")).lower()
        if expected_sha256:
            actual_sha256 = file_sha256(temp_path)
            if actual_sha256.lower() != expected_sha256:
                raise ValueError("Контрольная сумма обновления не совпала")

        return temp_path
    except Exception:
        try:
            os.remove(temp_path)
        except OSError:
            pass
        raise

def create_windows_updater(new_exe_path):
    if not getattr(sys, "frozen", False):
        raise RuntimeError("Автообновление доступно только в собранной Windows-версии приложения")
    if os.name != "nt":
        raise RuntimeError("Автообновление сейчас поддерживает только Windows exe")

    current_exe = sys.executable
    updater_path = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_updater_{os.getpid()}.bat")
    log_path = os.path.join(APP_DIR, f"{APP_NAME}_update.log")
    script = f"""@echo off
chcp 65001 >nul
set "APP={current_exe}"
set "NEW={new_exe_path}"
set "LOG={log_path}"
timeout /t 2 /nobreak >nul
for /l %%i in (1,1,60) do (
  copy /Y "%NEW%" "%APP%" >nul 2>nul
  if not errorlevel 1 (
    start "" "%APP%"
    del "%NEW%" >nul 2>nul
    del "%~f0" >nul 2>nul
    exit /b 0
  )
  timeout /t 1 /nobreak >nul
)
echo [%date% %time%] Не удалось заменить приложение >> "%LOG%"
start "" "%APP%"
exit /b 1
"""
    with open(updater_path, "w", encoding="utf-8") as updater_file:
        updater_file.write(script)
    return updater_path

def prepare_update_installer(update_info):
    new_exe_path = download_update_file(update_info)
    return create_windows_updater(new_exe_path)

def maybe_rename_windows_executable():
    if not getattr(sys, "frozen", False) or os.name != "nt":
        return False

    current_exe = os.path.abspath(sys.executable)
    target_exe = os.path.join(os.path.dirname(current_exe), APP_EXECUTABLE_NAME)
    if os.path.basename(current_exe).lower() == APP_EXECUTABLE_NAME.lower():
        return False

    updater_path = os.path.join(tempfile.gettempdir(), f"{APP_NAME}_rename_{os.getpid()}.bat")
    log_path = os.path.join(APP_DIR, f"{APP_NAME}_update.log")
    script = f"""@echo off
chcp 65001 >nul
set "OLD={current_exe}"
set "NEW={target_exe}"
set "LOG={log_path}"
timeout /t 1 /nobreak >nul
copy /Y "%OLD%" "%NEW%" >nul 2>nul
if errorlevel 1 (
  echo [%date% %time%] Не удалось создать "%NEW%" >> "%LOG%"
  start "" "%OLD%"
  del "%~f0" >nul 2>nul
  exit /b 1
)
start "" "%NEW%"
timeout /t 3 /nobreak >nul
del "%OLD%" >nul 2>nul
del "%~f0" >nul 2>nul
exit /b 0
"""
    with open(updater_path, "w", encoding="utf-8") as updater_file:
        updater_file.write(script)

    creationflags = subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0
    subprocess.Popen(["cmd", "/c", updater_path], creationflags=creationflags)
    return True

class ScanningApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"📦 {APP_NAME} — система учёта склада")
        self.configure(bg=BG_MAIN)
        self.geometry("1250x780")

        self.today_orders = []
        self.sheet = None
        self.all_existing_codes = set()
        self.current_legal_entity = None
        self.current_legal_entity_orders = []
        self.current_product_idx = 0
        self.current_order = None
        self.scanned_codes = []
        self.saved_codes_count = 0
        self.completed_orders = []
        self.current_legal_entity_products = []
        self.error_timer = None
        self.visible_order_groups = []
        self.current_group_key = None
        self.operation_in_progress = False
        self.update_required = False
        self.update_info = None
        self.telegram_poll_running = False
        self.last_sync_result = {"synced": 0, "failed": 0, "remaining": 0}
        self.product_catalog = load_product_catalog()
        os.makedirs(BACKUP_DIR, exist_ok=True)
        os.makedirs(REPORTS_DIR, exist_ok=True)

        self._build_ui()
        self.center_window()
        self.after(100, lambda: self.scan_entry.focus_set())
        self.after(150, lambda: self.refresh_from_sheet(initial=True))
        self.after(500, self.check_pending_prints)
        self.after(1200, self.check_for_updates)
        self.after(2500, self.sync_pending_telegram_async)
        self.after(4000, self.poll_telegram_bot_async)
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def load_data(self, show_empty_warning=True):
        self.today_orders, self.sheet, self.all_existing_codes = fetch_sheet_data()
        if show_empty_warning and not self.today_orders:
            messagebox.showwarning("Нет заданий",
                f"Нет заказов со статусом '{STATUS_NOT_COMPLETED}'.\n\n"
                f"Проверьте:\n"
                f"1. В таблице есть строки заказов\n"
                f"2. Колонка '{STATUS_COLUMN}' не заполнена как '{STATUS_COMPLETED}'\n"
                f"3. Колонка 'Отсканированные коды' заполнена не полностью")

    def apply_loaded_data(self, result, show_empty_warning):
        if len(result) == 4:
            self.today_orders, self.sheet, self.all_existing_codes, self.last_sync_result = result
        else:
            self.today_orders, self.sheet, self.all_existing_codes = result
            self.last_sync_result = {"synced": 0, "failed": 0, "remaining": len(load_pending_saves())}

        if show_empty_warning and not self.today_orders:
            messagebox.showwarning(
                "Нет заданий",
                f"Нет заказов со статусом '{STATUS_NOT_COMPLETED}'.\n\n"
                f"Проверьте:\n"
                f"1. В таблице есть строки заказов\n"
                f"2. Колонка '{STATUS_COLUMN}' не заполнена как '{STATUS_COMPLETED}'\n"
                f"3. Колонка 'Отсканированные коды' пустая или заполнена не полностью у активных заказов"
            )
        self.update_stats_display()

    def run_background(self, title, work, on_success=None, on_error=None, on_finally=None):
        def worker():
            try:
                result = work()
            except Exception as exc:
                logging.exception(title)

                def fail(exc=exc):
                    try:
                        if on_error:
                            on_error(exc)
                        else:
                            self.show_critical_error(title, exc)
                    finally:
                        if on_finally:
                            on_finally()

                try:
                    self.after(0, fail)
                except tk.TclError:
                    pass
                return

            def done():
                try:
                    if on_success:
                        on_success(result)
                finally:
                    if on_finally:
                        on_finally()

            try:
                self.after(0, done)
            except tk.TclError:
                pass

        threading.Thread(target=worker, daemon=True).start()

    def set_busy(self, message):
        self.operation_in_progress = True
        self.status_var.set(message)
        self.status_label.config(bg=BG_MAIN, fg=FG_MUTED)

    def clear_busy(self):
        self.operation_in_progress = False

    def ensure_update_allowed(self):
        if not self.update_required:
            return True
        self.show_error("Требуется обновить приложение перед работой")
        return False

    def apply_required_update_lock(self):
        self.status_var.set("⛔ Требуется обновление приложения")
        self.status_label.config(bg=ERROR_BG, fg=ERROR_FG)
        for button_name in (
            "refresh_btn",
            "import_btn",
            "catalog_btn",
            "control_btn",
            "select_btn",
            "undo_btn",
            "next_product_btn",
            "finish_btn",
            "report_btn",
        ):
            button = getattr(self, button_name, None)
            if button:
                button.config(state="disabled")

    def start_auto_update(self, update_info):
        self.status_var.set("⏳ Скачиваю обновление...")
        self.status_label.config(bg=BG_MAIN, fg=FG_MUTED)

        def worker():
            try:
                updater_path = prepare_update_installer(update_info)
            except Exception as exc:
                logging.exception("Не удалось подготовить автообновление")
                try:
                    self.after(0, lambda exc=exc: self.show_critical_error("Не удалось обновить приложение автоматически", exc))
                except tk.TclError:
                    pass
                return

            try:
                self.after(0, lambda: self.run_update_installer(updater_path))
            except tk.TclError:
                pass

        threading.Thread(target=worker, daemon=True).start()

    def run_update_installer(self, updater_path):
        self.status_var.set("⏳ Устанавливаю обновление...")
        self.status_label.config(bg=BG_MAIN, fg=FG_MUTED)
        creationflags = subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0
        subprocess.Popen(["cmd", "/c", updater_path], creationflags=creationflags)
        self.destroy()

    def send_telegram_documents_async(self, documents):
        documents = [
            (path, caption)
            for path, caption in documents
            if safe_telegram_document_path(path)
        ]
        if not documents:
            return

        def worker():
            for path, caption in documents:
                send_or_queue_telegram_document(path, caption)

        threading.Thread(target=worker, daemon=True).start()

    def sync_pending_telegram_async(self):
        def worker():
            result = sync_pending_telegram()
            if result.get("sent"):
                logging.info("Telegram: отправлено из очереди: %s", result["sent"])

        threading.Thread(target=worker, daemon=True).start()

    def send_telegram_alert_async(self, message, with_keyboard=True):
        if not telegram_is_configured():
            return

        def worker():
            ok, result = send_telegram_message(
                message,
                reply_markup=telegram_reports_keyboard() if with_keyboard else None,
            )
            if not ok:
                logging.warning("Telegram: сообщение не отправлено: %s", result)

        threading.Thread(target=worker, daemon=True).start()

    def log_duplicate_code_async(self, code):
        current_order = dict(self.current_order or {})

        def worker():
            details = []
            try:
                sheet = self.sheet
                if not sheet:
                    client = get_google_client()
                    sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)
                details = find_code_details_in_sheet(sheet, code)
            except Exception:
                logging.exception("Не удалось определить строку дублирующегося КИЗа")
            details.extend(find_code_details_in_pending_saves(code))

            detail_text = format_duplicate_code_details(code, details, current_order=current_order)
            logging.warning("%s", detail_text)
            if telegram_is_configured():
                ok, result = send_telegram_message(
                    f"{APP_NAME}: найден дублирующийся КИЗ\n\n" + detail_text,
                    reply_markup=telegram_reports_keyboard(),
                )
                if not ok:
                    logging.warning("Telegram: дубль КИЗ не отправлен: %s", result)

        threading.Thread(target=worker, daemon=True).start()

    def get_sheet_for_telegram(self):
        if self.sheet:
            return self.sheet
        client = get_google_client()
        sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)
        ensure_import_sheet_layout(sheet)
        self.sheet = sheet
        return sheet

    def send_today_scan_report_to_chat(self, chat_id, token):
        sheet = self.get_sheet_for_telegram()
        result = create_day_report_excel(sheet)
        if result.get("empty"):
            send_telegram_message_to_chat(
                chat_id,
                "За сегодня пока нет отсканированных КИЗов для отчёта.",
                token,
                reply_markup=telegram_reports_keyboard(),
            )
            return

        caption = (
            f"{APP_NAME}: сканы за сегодня на текущий момент\n"
            f"Всего КИЗов: {result['total_report_rows']}\n"
            f"Терминал: {result['terminal_count']}; "
            f"Перечисление: {result['transfer_count']}; "
            f"Не распознано: {result['unknown_count']}"
        )
        send_telegram_document_to_chat(result["filename"], chat_id, caption, token)

    def send_today_log_to_chat(self, chat_id, token):
        log_path = create_today_log_file()
        send_telegram_document_to_chat(
            log_path,
            chat_id,
            f"{APP_NAME}: сегодняшний лог по времени и ошибкам",
            token,
        )

    def send_document_list_to_chat(self, chat_id, token):
        sheet = self.get_sheet_for_telegram()
        documents = build_document_summaries_from_gsheet(sheet, limit=12)
        if not documents:
            send_telegram_message_to_chat(
                chat_id,
                "Импортированные документы в Google Sheets не найдены.",
                token,
                reply_markup=telegram_reports_keyboard(),
            )
            return

        lines = [f"{APP_NAME}: документы по импорту"]
        for idx, document in enumerate(documents, start=1):
            lines.append(
                f"{idx}. {document['source_file']} - "
                f"{document['scanned_blocks']}/{document['plan_blocks']} КИЗ, "
                f"позиций {document['completed_positions']}/{document['positions']}"
            )
        send_telegram_message_to_chat(
            chat_id,
            "\n".join(lines),
            token,
            reply_markup=telegram_documents_keyboard(documents),
        )

    def send_document_report_to_chat(self, chat_id, token, document_key):
        sheet = self.get_sheet_for_telegram()
        result = create_document_report_excel(sheet, document_key)
        if result.get("empty"):
            send_telegram_message_to_chat(
                chat_id,
                "Документ не найден. Обновите список документов в боте.",
                token,
                reply_markup=telegram_reports_keyboard(),
            )
            return

        caption = (
            f"{APP_NAME}: документ {result['source_file']}\n"
            f"КИЗ: {result['scanned_blocks']}/{result['plan_blocks']}\n"
            f"Позиций: {result['completed_positions']}/{result['positions']}\n"
            f"Осталось КИЗ: {result['remaining_blocks']}"
        )
        if result.get("pending_blocks"):
            caption += f"\nВ локальной очереди: {result['pending_blocks']}"
        send_telegram_document_to_chat(result["filename"], chat_id, caption, token)

    def send_telegram_menu_to_chat(self, chat_id, token):
        send_telegram_message_to_chat(
            chat_id,
            f"{APP_NAME}: выберите файл, который нужно получить, или отправьте Excel-файл для импорта.",
            token,
            reply_markup=telegram_reports_keyboard(),
        )

    def start_telegram_import_ui(self, file_name):
        self.status_var.set(f"⏳ Импортирую Excel из Telegram: {file_name}")
        self.status_label.config(bg=BG_MAIN, fg=FG_MUTED)
        self.import_btn.config(state="disabled")
        self.refresh_btn.config(state="disabled")

    def finish_telegram_import_ui(self, status_message, loaded=None):
        try:
            if loaded is not None:
                self.product_catalog = load_product_catalog()
                self.apply_loaded_data(loaded, show_empty_warning=False)
                self.reset_current_selection()
                self.refresh_legal_list()
            self.status_var.set(status_message)
            self.status_label.config(bg=BG_MAIN, fg=FG_MUTED)
        except Exception:
            logging.exception("Telegram: не удалось обновить интерфейс после импорта")
            self.status_var.set("Excel импортирован из Telegram, обновите список вручную")
        finally:
            self.operation_in_progress = False
            if not self.update_required:
                self.import_btn.config(state="normal")
                self.refresh_btn.config(state="normal")

    def handle_telegram_document_message(self, document, chat_id, token):
        chat_id = normalize_text(chat_id)
        file_name = telegram_document_file_name(document)

        def safe_send(text, reply_markup=None):
            try:
                send_telegram_message_to_chat(chat_id, text, token, reply_markup=reply_markup)
            except Exception:
                logging.exception("Telegram: не удалось отправить ответ по импорту Excel")

        if not telegram_document_is_supported_excel(document):
            safe_send(
                "Файл не импортирован.\n\n"
                "Отправьте Excel-файл в формате .xlsx или .xlsm.",
                reply_markup=telegram_reports_keyboard(),
            )
            return

        if self.update_required:
            safe_send(
                f"Файл не импортирован: сначала нужно обновить {APP_NAME} на компьютере склада.",
                reply_markup=telegram_reports_keyboard(),
            )
            return

        if self.operation_in_progress:
            safe_send(
                f"Файл не импортирован: {APP_NAME} сейчас занят другой операцией. "
                "Отправьте Excel-файл повторно после завершения операции.",
                reply_markup=telegram_reports_keyboard(),
            )
            return

        self.operation_in_progress = True
        temp_path = None
        finish_scheduled = False

        def schedule_finish(status_message, loaded=None):
            nonlocal finish_scheduled
            if finish_scheduled:
                return
            finish_scheduled = True
            try:
                self.after(0, lambda: self.finish_telegram_import_ui(status_message, loaded=loaded))
            except tk.TclError:
                self.operation_in_progress = False

        try:
            try:
                self.after(0, lambda: self.start_telegram_import_ui(file_name))
            except tk.TclError:
                pass

            safe_send(f"Получил Excel-файл: {file_name}\nНачинаю импорт в Google Sheets.")

            temp_path, source_file_name = download_telegram_document_to_temp(token, document)
            file_hash = file_sha256(temp_path)
            previous_import = find_successful_import_by_file_hash(file_hash)
            if previous_import:
                raw_sources = previous_import.get("sources", [])
                if isinstance(raw_sources, str):
                    raw_sources = [raw_sources]
                elif not isinstance(raw_sources, list):
                    raw_sources = [raw_sources]
                previous_sources = ", ".join(normalize_text(source) for source in raw_sources[:3] if normalize_text(source))
                previous_date = normalize_text(previous_import.get("timestamp")) or "дата неизвестна"
                details = [
                    "Повторный импорт заблокирован.",
                    "",
                    f"Файл: {file_name}",
                    f"Уже импортирован: {previous_date}",
                ]
                if previous_sources:
                    details.append(f"В истории: {previous_sources}")
                safe_send("\n".join(details), reply_markup=telegram_reports_keyboard())
                schedule_finish(f"Повторный Excel из Telegram заблокирован: {file_name}")
                return

            preview = prepare_excel_import([temp_path], source_names={temp_path: source_file_name})
            errors = preview.get("errors", [])
            warnings = preview.get("warnings", [])
            new_records = preview.get("new_records", [])
            duplicate_records = preview.get("duplicate_records", [])

            if not new_records:
                lines = [
                    "Новых позиций для импорта не найдено.",
                    "",
                    f"Файл: {file_name}",
                    f"Строк в файле: {preview.get('source_rows_count', 0)}",
                    f"Повторных позиций: {len(duplicate_records)}",
                    f"Адресов получено из координат: {preview.get('geocoded_count', 0)}",
                    f"Координат без адреса: {preview.get('geocode_failed_count', 0)}",
                ]
                if duplicate_records and not errors:
                    lines.insert(0, "Файл уже загружен в Google Sheets, повторный импорт заблокирован.")
                if errors:
                    lines.extend(["", "Ошибки:", "\n".join(errors[:5])])
                if warnings:
                    lines.extend(["", "Предупреждения:", "\n".join(warnings[:5])])
                safe_send("\n".join(lines), reply_markup=telegram_reports_keyboard())
                schedule_finish(f"Excel из Telegram не содержит новых позиций: {file_name}")
                return

            import_result = append_import_records(new_records)
            imported_count = import_result.get("imported", 0)
            if imported_count <= 0:
                safe_send(
                    "Новые позиции не были добавлены: все строки уже есть в Google Sheets.",
                    reply_markup=telegram_reports_keyboard(),
                )
                schedule_finish(f"Excel из Telegram не добавил новых позиций: {file_name}")
                return

            loaded = None
            refresh_note = ""
            try:
                loaded = fetch_sheet_data_with_sync()
            except Exception:
                logging.exception("Telegram: Excel импортирован, но список заказов не обновился")
                refresh_note = f"Список в окне {APP_NAME} не обновился автоматически. Нажмите «Обновить»."

            imported_blocks = sum(parse_int_value(record.get("Кол-во блок")) for record in new_records)
            lines = [
                f"{APP_NAME}: Excel импортирован из Telegram",
                "",
                f"Документ: {file_name}",
                f"Позиций загружено: {imported_count}",
                f"Повторно пропущено: {import_result.get('duplicates', 0)}",
                f"План КИЗ: {imported_blocks}",
                f"Адресов получено из координат: {preview.get('geocoded_count', 0)}",
                f"Координат без адреса: {preview.get('geocode_failed_count', 0)}",
                "",
                "Документ доступен в разделе «Документы по импорту».",
            ]
            if warnings:
                lines.extend(["", "Предупреждения:", "\n".join(warnings[:5])])
            if errors:
                lines.extend(["", "Ошибки в отдельных строках:", "\n".join(errors[:5])])
            if refresh_note:
                lines.extend(["", refresh_note])
            safe_send("\n".join(lines), reply_markup=telegram_reports_keyboard())
            schedule_finish(f"✅ Excel импортирован из Telegram: {file_name}", loaded=loaded)
        except Exception as exc:
            logging.exception("Telegram: не удалось импортировать Excel-файл")
            safe_send(
                "Не удалось импортировать Excel-файл.\n\n"
                f"Файл: {file_name}\n"
                f"Причина: {exc}\n\n"
                f"Подробности записаны в лог {APP_NAME}.",
                reply_markup=telegram_reports_keyboard(),
            )
            try:
                log_path = create_today_log_file()
                send_telegram_document_to_chat(log_path, chat_id, f"{APP_NAME}: лог ошибки импорта Excel", token)
            except Exception:
                logging.exception("Telegram: не удалось отправить лог ошибки импорта")
            schedule_finish(f"Ошибка импорта Excel из Telegram: {file_name}")
        finally:
            if temp_path:
                try:
                    os.remove(temp_path)
                except OSError:
                    pass
            if not finish_scheduled:
                schedule_finish(f"Импорт Excel из Telegram завершён: {file_name}")

    def handle_telegram_message(self, message, settings, token):
        chat = message.get("chat") or {}
        sender = message.get("from") or {}
        chat_id = normalize_text(chat.get("id"))
        sender_id = normalize_text(sender.get("id"))
        if not (telegram_chat_is_authorized(chat_id, settings) or telegram_chat_is_authorized(sender_id, settings)):
            logging.warning("Telegram: отказано неизвестному chat_id=%s sender_id=%s", chat_id, sender_id)
            return

        if message.get("document"):
            self.handle_telegram_document_message(message["document"], chat_id or sender_id, token)
            return

        self.send_telegram_menu_to_chat(chat_id or sender_id, token)

    def handle_telegram_callback(self, callback_query, settings, token):
        query_id = callback_query.get("id")
        sender = callback_query.get("from") or {}
        message = callback_query.get("message") or {}
        chat = message.get("chat") or {}
        chat_id = normalize_text(chat.get("id") or sender.get("id"))
        sender_id = normalize_text(sender.get("id"))

        if not (telegram_chat_is_authorized(chat_id, settings) or telegram_chat_is_authorized(sender_id, settings)):
            try:
                answer_telegram_callback_query(token, query_id, "Нет доступа")
            except Exception:
                logging.exception("Telegram: не удалось ответить на запрещенный callback")
            logging.warning("Telegram: запрещенный callback chat_id=%s sender_id=%s", chat_id, sender_id)
            return

        data = normalize_text(callback_query.get("data"))
        try:
            if data == TELEGRAM_CALLBACK_TODAY_SCANS:
                answer_telegram_callback_query(token, query_id, "Готовлю отчёт...")
                self.send_today_scan_report_to_chat(chat_id, token)
            elif data == TELEGRAM_CALLBACK_DOCUMENTS:
                answer_telegram_callback_query(token, query_id, "Открываю документы...")
                self.send_document_list_to_chat(chat_id, token)
            elif data.startswith(TELEGRAM_CALLBACK_DOCUMENT_PREFIX):
                answer_telegram_callback_query(token, query_id, "Готовлю документ...")
                self.send_document_report_to_chat(
                    chat_id,
                    token,
                    data[len(TELEGRAM_CALLBACK_DOCUMENT_PREFIX):],
                )
            elif data == TELEGRAM_CALLBACK_TODAY_LOG:
                answer_telegram_callback_query(token, query_id, "Готовлю лог...")
                self.send_today_log_to_chat(chat_id, token)
            else:
                answer_telegram_callback_query(token, query_id, "Открываю меню")
                self.send_telegram_menu_to_chat(chat_id, token)
        except Exception as exc:
            logging.exception("Telegram: не удалось обработать команду")
            try:
                send_telegram_message_to_chat(
                    chat_id,
                    f"Не удалось выполнить команду: {exc}",
                    token,
                    reply_markup=telegram_reports_keyboard(),
                )
            except Exception:
                logging.exception("Telegram: не удалось отправить ошибку команды")

    def process_telegram_updates(self, settings):
        token = normalize_text(settings.get("bot_token"))
        state = load_telegram_state()
        last_update_id = parse_int_value(state.get("last_update_id"))
        updates = fetch_telegram_updates(token, offset=last_update_id + 1 if last_update_id else None)
        max_update_id = last_update_id

        for update in updates:
            update_id = parse_int_value(update.get("update_id"))
            if update_id:
                max_update_id = max(max_update_id, update_id)
            if update.get("callback_query"):
                self.handle_telegram_callback(update["callback_query"], settings, token)
            elif update.get("message"):
                self.handle_telegram_message(update["message"], settings, token)

        if max_update_id != last_update_id:
            state["last_update_id"] = max_update_id
            state["updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            save_telegram_state(state)

    def poll_telegram_bot_async(self):
        settings = load_telegram_settings()
        delay_ms = 5000 if telegram_is_configured(settings) else 15000
        if self.telegram_poll_running:
            self.after(delay_ms, self.poll_telegram_bot_async)
            return
        if not telegram_is_configured(settings):
            self.after(delay_ms, self.poll_telegram_bot_async)
            return

        self.telegram_poll_running = True

        def worker():
            try:
                self.process_telegram_updates(settings)
            except Exception:
                logging.exception("Telegram: не удалось проверить команды бота")
            finally:
                def finish():
                    self.telegram_poll_running = False
                    self.after(delay_ms, self.poll_telegram_bot_async)

                try:
                    self.after(0, finish)
                except tk.TclError:
                    pass

        threading.Thread(target=worker, daemon=True).start()

    def check_for_updates(self):
        if not UPDATE_INFO_URL:
            return

        def worker():
            try:
                update_info = fetch_update_info()
            except Exception as exc:
                logging.info("Не удалось проверить обновления: %s", exc)
                return

            try:
                self.after(0, lambda: self.handle_update_info(update_info))
            except tk.TclError:
                pass

        threading.Thread(target=worker, daemon=True).start()

    def handle_update_info(self, update_info):
        if not update_info:
            return

        latest_version = normalize_text(update_info.get("latest_version"))
        min_supported_version = normalize_text(update_info.get("min_supported_version"))
        message = normalize_text(update_info.get("message"))
        update_available = bool(latest_version) and compare_versions(APP_VERSION, latest_version) < 0
        below_min_version = bool(min_supported_version) and compare_versions(APP_VERSION, min_supported_version) < 0

        if not update_available and not below_min_version:
            return

        self.update_info = update_info
        self.update_required = True
        self.apply_required_update_lock()

        lines = [
            "Вышла новая версия приложения.",
            "Программа скачает обновление, перезапустится и продолжит работу уже в новой версии.",
            "",
            f"Текущая версия: {APP_VERSION}",
            f"Новая версия: {latest_version or 'не указана'}",
        ]
        if message:
            lines.extend(["", message])
        self.status_var.set("⏳ Найдено обновление, начинаю установку...")
        logging.info("Запущено автообновление: %s -> %s", APP_VERSION, latest_version)
        messagebox.showinfo("Обновление приложения", "\n".join(lines))
        self.start_auto_update(update_info)

    def center_window(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth() - self.winfo_width()) // 2
        y = (self.winfo_screenheight() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

    def show_error(self, message, popup=True):
        logging.warning("Ошибка для пользователя: %s", message)
        self.status_var.set(f"❌ {message}")
        self.status_label.config(bg=ERROR_BG, fg=ERROR_FG)
        if self.error_timer:
            self.after_cancel(self.error_timer)
        self.error_timer = self.after(3000, self.clear_error)
        if popup:
            messagebox.showerror(
                "Ошибка",
                f"Причина: {message}\n\nЕсли ошибка повторяется, подробности будут в логе:\n{LOG_FILE}"
            )

    def show_critical_error(self, title, exc_or_message):
        if isinstance(exc_or_message, BaseException):
            message = str(exc_or_message)
            logging.error(
                title,
                exc_info=(type(exc_or_message), exc_or_message, exc_or_message.__traceback__)
            )
            detail = format_exception_message(title, exc_or_message)
        else:
            message = str(exc_or_message)
            logging.error("%s: %s", title, message)
            detail = f"{title}\n\nПричина: {message}\n\nПодробности записаны в лог:\n{LOG_FILE}"

        self.show_error(message, popup=False)
        messagebox.showerror("Ошибка", detail)
        self.send_telegram_alert_async(f"{APP_NAME}: ошибка приложения\n\n" + detail[:3800])
        self.send_telegram_documents_async(collect_operational_documents(include_error_log=True))

    def report_callback_exception(self, exc_type, exc_value, exc_traceback):
        logging.error(
            "Ошибка в интерфейсе",
            exc_info=(exc_type, exc_value, exc_traceback)
        )
        try:
            self.show_error(str(exc_value), popup=False)
            detail = format_exception_message("Ошибка в интерфейсе", exc_value)
            messagebox.showerror(
                "Ошибка",
                detail
            )
            self.send_telegram_alert_async(f"{APP_NAME}: ошибка интерфейса\n\n" + detail[:3800])
            self.send_telegram_documents_async(collect_operational_documents(include_error_log=True))
        except Exception:
            pass

    def clear_error(self):
        if self.update_required:
            self.status_var.set("⛔ Требуется обновление приложения")
            self.status_label.config(bg=ERROR_BG, fg=ERROR_FG)
            self.error_timer = None
            return
        self.status_var.set("✅ Готов к работе")
        self.status_label.config(bg=BG_MAIN, fg=FG_MUTED)
        self.error_timer = None

    def validate_code(self, code):
        if not code:
            return False, "Код пустой"
        if not code.startswith('01'):
            return False, "КИЗ должен начинаться с 01"
        if len(code) < KIZ_MIN_LENGTH:
            return False, f"Код слишком короткий для КИЗа (минимум {KIZ_MIN_LENGTH} символов)"
        if len(code) > KIZ_MAX_LENGTH:
            return False, f"Код слишком длинный для КИЗа (максимум {KIZ_MAX_LENGTH} символов)"
        if re.search(r'[а-яА-ЯёЁ]', code):
            return False, "Код содержит русские буквы! Используйте только латиницу"
        if re.search(r'\s', code):
            return False, "Код содержит пробелы или переносы"
        if not re.fullmatch(r'[\x1d\x21-\x7E]+', code):
            return False, "Код содержит недопустимые символы"
        return True, ""

    def undo_last_scan(self):
        if not self.ensure_update_allowed():
            return

        if self.operation_in_progress:
            self.show_error("Дождитесь завершения текущей операции")
            return

        if not self.current_order:
            self.show_error("Нет активной позиции")
            return

        if not self.scanned_codes:
            self.show_error("Нет кодов для отмены")
            return

        if len(self.scanned_codes) <= self.saved_codes_count:
            self.show_error("Нельзя отменить коды, уже записанные в Google Sheets")
            return

        removed_code = self.scanned_codes.pop()
        self.all_existing_codes.discard(removed_code)
        write_scan_backup("undo_scan", self.current_order, code=removed_code, codes=self.scanned_codes.copy())

        plan_blocks = get_plan_blocks(self.current_order)

        scanned_count = len(self.scanned_codes)
        self.progress_label.config(text=f"{scanned_count} / {plan_blocks}")
        self.last_code_label.config(text=f"Отменён код: {removed_code[:40]}...")
        self.status_var.set(f"↩️ Отменён последний код ({scanned_count}/{plan_blocks})")

        if scanned_count < plan_blocks:
            self.next_product_btn.config(state="disabled")
            self.finish_btn.config(state="normal")

        self.scan_entry.focus_set()

    def _build_ui(self):
        main = tk.Frame(self, bg=BG_MAIN)
        main.pack(fill="both", expand=True, padx=25, pady=20)

        title = tk.Label(main, text="📦 УЧЁТ СКАНИРОВАНИЯ БЛОКОВ",
                        bg=BG_MAIN, fg=FG_TEXT, font=("Segoe UI", 24, "bold"))
        title.pack(pady=(0, 5))

        date_label = tk.Label(main, text=f"Дата: {datetime.now().strftime('%d.%m.%Y')}",
                             bg=BG_MAIN, fg=FG_MUTED, font=("Segoe UI", 12))
        date_label.pack(pady=(0, 20))

        content = tk.Frame(main, bg=BG_MAIN)
        content.pack(fill="both", expand=True)

        left_panel = tk.Frame(content, bg=BG_MAIN)
        left_panel.pack(side="left", fill="both", expand=True, padx=(0, 15))

        list_card = tk.Frame(left_panel, bg=BG_CARD, relief="flat", bd=1, highlightbackground=BORDER)
        list_card.pack(fill="both", expand=True)

        list_header = tk.Frame(list_card, bg=BG_CARD)
        list_header.pack(fill="x", padx=20, pady=(15, 10))

        tk.Label(list_header, text="🏢 Заказы на сегодня",
                bg=BG_CARD, fg=FG_TEXT, font=("Segoe UI", 14, "bold")).pack(side="left")

        self.refresh_btn = AppButton(list_header, text="🔄 ОБНОВИТЬ",
                                     bg=INFO, fg="white", font=("Segoe UI", 9, "bold"),
                                     command=self.refresh_from_sheet, relief="flat", cursor="hand2")
        self.refresh_btn.pack(side="right")

        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self.refresh_legal_list())
        self.search_entry = tk.Entry(list_card, textvariable=self.search_var, bg=BG_MAIN, fg=FG_TEXT,
                                     font=("Segoe UI", 11), relief="flat", bd=0,
                                     highlightbackground=BORDER, highlightcolor=ACCENT,
                                     highlightthickness=1, insertbackground=FG_TEXT)
        self.search_entry.pack(fill="x", padx=15, pady=(0, 10))

        tools_frame = tk.Frame(list_card, bg=BG_CARD)
        tools_frame.pack(fill="x", padx=15, pady=(0, 10))

        self.import_btn = AppButton(
            tools_frame,
            text="📥 ИМПОРТ EXCEL",
            bg=SUCCESS,
            fg="white",
            font=("Segoe UI", 9, "bold"),
            command=self.import_excel_orders,
            relief="flat",
            cursor="hand2"
        )
        self.import_btn.pack(side="left", fill="x", expand=True, padx=(0, 6))

        self.catalog_btn = AppButton(
            tools_frame,
            text="📦 ТОВАРЫ",
            bg=WARNING,
            fg="white",
            font=("Segoe UI", 9, "bold"),
            command=self.show_product_catalog,
            relief="flat",
            cursor="hand2"
        )
        self.catalog_btn.pack(side="left", fill="x", expand=True, padx=(0, 6))

        self.control_btn = AppButton(
            tools_frame,
            text="📊 КОНТРОЛЬ",
            bg=INFO,
            fg="white",
            font=("Segoe UI", 9, "bold"),
            command=self.show_control_panel,
            relief="flat",
            cursor="hand2"
        )
        self.control_btn.pack(side="left", fill="x", expand=True)

        list_container = tk.Frame(list_card, bg=BG_CARD)
        list_container.pack(fill="both", expand=True, padx=15, pady=(0, 15))

        self.legal_listbox = tk.Listbox(list_container, bg=BG_CARD, fg=FG_TEXT,
                                        font=("Segoe UI", 11), selectmode=tk.SINGLE,
                                        relief="flat", selectbackground=ACCENT, selectforeground="white")
        self.legal_listbox.pack(side="left", fill="both", expand=True)

        scrollbar = tk.Scrollbar(list_container, orient="vertical", command=self.legal_listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.legal_listbox.config(yscrollcommand=scrollbar.set)

        self.refresh_legal_list()

        self.select_btn = AppButton(left_panel, text="✅ ВЫБРАТЬ ЗАКАЗ",
                                   bg=ACCENT, fg="white", font=("Segoe UI", 12, "bold"),
                                   command=self.select_legal_entity, relief="flat", pady=12,
                                   cursor="hand2")
        self.select_btn.pack(pady=(15, 0), fill="x")

        right_panel = tk.Frame(content, bg=BG_MAIN)
        right_panel.pack(side="right", fill="both", expand=True, padx=(15, 0))

        info_card = tk.Frame(right_panel, bg=BG_CARD, relief="flat", bd=1, highlightbackground=BORDER)
        info_card.pack(fill="x", pady=(0, 15))

        tk.Label(info_card, text="📋 ТЕКУЩАЯ ПОЗИЦИЯ",
                bg=BG_CARD, fg=ACCENT, font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=20, pady=(15, 10))

        self.current_info = tk.Label(info_card, text="Не выбрано",
                                    bg=BG_CARD, fg=FG_TEXT, font=("Segoe UI", 11),
                                    wraplength=400, justify="left")
        self.current_info.pack(anchor="w", padx=20, pady=(0, 10))

        positions_frame = tk.Frame(info_card, bg=BG_CARD)
        positions_frame.pack(fill="x", padx=20, pady=(0, 10))

        self.position_label = tk.Label(positions_frame, text="", bg=BG_CARD, fg=WARNING, font=("Segoe UI", 11, "bold"))
        self.position_label.pack(side="left")

        progress_frame = tk.Frame(info_card, bg=BG_CARD)
        progress_frame.pack(fill="x", padx=20, pady=(0, 15))

        tk.Label(progress_frame, text="Сканирование:", bg=BG_CARD, fg=FG_MUTED, font=("Segoe UI", 11)).pack(side="left")
        self.progress_label = tk.Label(progress_frame, text="0 / 0", bg=BG_CARD, fg=SUCCESS, font=("Segoe UI", 14, "bold"))
        self.progress_label.pack(side="left", padx=(10, 0))

        scan_card = tk.Frame(right_panel, bg=BG_CARD, relief="flat", bd=1, highlightbackground=BORDER)
        scan_card.pack(fill="x", pady=(0, 15))

        tk.Label(scan_card, text="🔍 СКАНИРОВАНИЕ КОДА",
                bg=BG_CARD, fg=ACCENT, font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=20, pady=(15, 10))

        self.scan_entry = tk.Entry(scan_card, bg=BG_MAIN, fg=FG_TEXT, font=("Segoe UI", 14),
                                   relief="flat", bd=0, highlightbackground=BORDER,
                                   highlightcolor=ACCENT, highlightthickness=1,
                                   insertbackground=FG_TEXT)
        self.scan_entry.pack(fill="x", padx=20, pady=(0, 10))
        self.scan_entry.bind("<Return>", self.on_scan)

        self.last_code_label = tk.Label(scan_card, text="", bg=BG_CARD, fg=SUCCESS, font=("Segoe UI", 10))
        self.last_code_label.pack(anchor="w", padx=20, pady=(5, 5))

        actions_frame = tk.Frame(right_panel, bg=BG_MAIN)
        actions_frame.pack(fill="x", pady=(0, 15))

        self.undo_btn = AppButton(actions_frame, text="↩️ ОТМЕНИТЬ ПОСЛЕДНИЙ КОД",
                                 bg=DANGER, fg="white", font=("Segoe UI", 10, "bold"),
                                 command=self.undo_last_scan, relief="flat", state="disabled",
                                 cursor="hand2")
        self.undo_btn.pack(side="left", fill="x", expand=True, padx=(0, 10), pady=5)

        self.next_product_btn = AppButton(actions_frame, text="➡️ СЛЕДУЮЩАЯ ПОЗИЦИЯ",
                                         bg=WARNING, fg="white", font=("Segoe UI", 11, "bold"),
                                         command=self.next_product, relief="flat", state="disabled",
                                         cursor="hand2")
        self.next_product_btn.pack(side="left", fill="x", expand=True, padx=(0, 10), pady=5)

        self.finish_btn = AppButton(actions_frame, text="🏁 ЗАВЕРШИТЬ ЗАКАЗ",
                                   bg=SUCCESS, fg="white", font=("Segoe UI", 11, "bold"),
                                   command=self.finish_legal_entity, relief="flat", state="disabled",
                                   cursor="hand2")
        self.finish_btn.pack(side="right", fill="x", expand=True, padx=(10, 0), pady=5)

        stats_card = tk.Frame(right_panel, bg=BG_CARD, relief="flat", bd=1, highlightbackground=BORDER)
        stats_card.pack(fill="x")

        tk.Label(stats_card, text="📊 СТАТИСТИКА",
                bg=BG_CARD, fg=ACCENT, font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=20, pady=(15, 10))

        stats_frame = tk.Frame(stats_card, bg=BG_CARD)
        stats_frame.pack(fill="x", padx=20, pady=(0, 15))

        self.completed_count_label = tk.Label(stats_frame, text="Выполнено: 0", bg=BG_CARD, fg=FG_TEXT, font=("Segoe UI", 11))
        self.completed_count_label.pack(side="left", padx=(0, 20))

        self.total_blocks_label = tk.Label(stats_frame, text="Блоков: 0", bg=BG_CARD, fg=FG_TEXT, font=("Segoe UI", 11))
        self.total_blocks_label.pack(side="left")

        stats_frame_2 = tk.Frame(stats_card, bg=BG_CARD)
        stats_frame_2.pack(fill="x", padx=20, pady=(0, 15))

        self.active_orders_label = tk.Label(stats_frame_2, text="Активных заказов: 0", bg=BG_CARD, fg=FG_TEXT, font=("Segoe UI", 10))
        self.active_orders_label.pack(side="left", padx=(0, 20))

        self.pending_saves_label = tk.Label(stats_frame_2, text="Очередь записи: 0", bg=BG_CARD, fg=FG_TEXT, font=("Segoe UI", 10))
        self.pending_saves_label.pack(side="left")

        self.report_btn = AppButton(right_panel, text="📊 ЗАВЕРШИТЬ ДЕНЬ (ОТЧЁТ)",
                                   bg=INFO, fg="white", font=("Segoe UI", 11, "bold"),
                                   command=self.end_day, relief="flat", pady=10,
                                   cursor="hand2")
        self.report_btn.pack(fill="x", pady=(10, 0))

        status_frame = tk.Frame(main, bg=BG_MAIN)
        status_frame.pack(fill="x", pady=(20, 0))

        self.status_var = tk.StringVar(value="✅ Готов к работе")
        self.status_label = tk.Label(status_frame, textvariable=self.status_var,
                                     bg=BG_MAIN, fg=FG_MUTED, font=("Segoe UI", 10))
        self.status_label.pack()

        version_frame = tk.Frame(main, bg=BG_MAIN)
        version_frame.pack(fill="x", pady=(6, 0))
        tk.Label(
            version_frame,
            text=f"Версия: {APP_VERSION}",
            bg=BG_MAIN,
            fg=FG_MUTED,
            font=("Segoe UI", 9),
        ).pack(side="left")

    def refresh_legal_list(self):
        self.legal_listbox.delete(0, tk.END)
        self.visible_order_groups = []
        grouped_orders = {}
        search_text = normalize_text(self.search_var.get()).lower() if hasattr(self, "search_var") else ""

        for order in self.today_orders:
            key = order_group_key(order)
            client, payment_type, address = key
            client = client or "Клиент не указан"
            payment_type = payment_type or "Оплата не указана"
            address = address or "Адрес не указан"
            search_area = " ".join([
                client,
                payment_type,
                address,
                normalize_text(order.get("Торговый представитель")),
                normalize_text(order.get("Товары")),
            ]).lower()
            if search_text and search_text not in search_area:
                continue
            grouped_orders.setdefault((client, payment_type, address), []).append(order)

        for key in sorted(grouped_orders.keys(), key=lambda item: (item[0].lower(), item[1].lower(), item[2].lower())):
            client, payment_type, address = key
            count = len(grouped_orders[key])
            self.visible_order_groups.append(key)
            self.legal_listbox.insert(tk.END, f"{client} | {payment_type} | {count} поз. | {address}")
        self.update_stats_display()

    def reset_current_selection(self):
        self.current_legal_entity = None
        self.current_group_key = None
        self.current_legal_entity_orders = []
        self.current_product_idx = 0
        self.current_order = None
        self.scanned_codes = []
        self.saved_codes_count = 0
        self.current_legal_entity_products = []
        self.current_info.config(text="Не выбрано")
        self.position_label.config(text="")
        self.progress_label.config(text="0 / 0")
        self.next_product_btn.config(state="disabled")
        self.finish_btn.config(state="disabled")
        self.undo_btn.config(state="disabled")
        self.last_code_label.config(text="")

    def refresh_from_sheet(self, initial=False):
        if not self.ensure_update_allowed():
            return

        if self.operation_in_progress:
            self.show_error("Дождитесь завершения текущей операции")
            return

        if self.current_order and not initial:
            if not messagebox.askyesno(
                "Обновить список?",
                "Есть выбранный заказ. Обновление сбросит текущий выбор и несохранённые сканы.\n\nПродолжить?"
            ):
                return

        self.set_busy("⏳ Обновляю список заказов...")
        self.refresh_btn.config(state="disabled")
        self.import_btn.config(state="disabled")

        def on_success(result):
            self.apply_loaded_data(result, show_empty_warning=initial)
            self.reset_current_selection()
            self.refresh_legal_list()
            sync_result = self.last_sync_result or {}
            if sync_result.get("synced"):
                self.status_var.set(f"✅ Список обновлён, отправлено из очереди: {sync_result['synced']}")
            else:
                self.status_var.set("✅ Список заказов обновлён")
            self.status_label.config(bg=BG_MAIN, fg=FG_MUTED)

        def on_error(exc):
            self.show_critical_error("Не удалось обновить список заказов", exc)

        def on_finally():
            self.refresh_btn.config(state="normal")
            self.import_btn.config(state="normal")
            self.clear_busy()

        self.run_background(
            "Не удалось обновить список заказов",
            fetch_sheet_data_with_sync,
            on_success=on_success,
            on_error=on_error,
            on_finally=on_finally
        )

    def import_excel_orders(self):
        if not self.ensure_update_allowed():
            return

        if self.operation_in_progress:
            self.show_error("Дождитесь завершения текущей операции")
            return

        file_paths = filedialog.askopenfilenames(
            title="Выберите Excel-файлы заказов",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm"),
                ("All files", "*.*"),
            ]
        )
        if not file_paths:
            return

        self.set_busy("⏳ Проверяю Excel-файлы перед импортом...")
        self.import_btn.config(state="disabled")
        self.refresh_btn.config(state="disabled")

        def work():
            return prepare_excel_import(list(file_paths))

        def on_success(preview):
            self.clear_busy()
            self.import_btn.config(state="normal")
            self.refresh_btn.config(state="normal")

            errors = preview.get("errors", [])
            warnings = preview.get("warnings", [])
            new_records = preview.get("new_records", [])
            duplicate_records = preview.get("duplicate_records", [])

            if not new_records:
                details = [
                    f"Файлов проверено: {preview.get('files_count', 0)}",
                    f"Строк в файлах: {preview.get('source_rows_count', 0)}",
                    f"Адресов получено из координат: {preview.get('geocoded_count', 0)}",
                    f"Координат без адреса: {preview.get('geocode_failed_count', 0)}",
                    f"Дублей найдено: {len(duplicate_records)}",
                ]
                if errors:
                    details.append("\nОшибки:\n" + "\n".join(errors[:6]))
                if warnings:
                    details.append("\nПредупреждения:\n" + "\n".join(warnings[:6]))
                messagebox.showwarning("Импорт Excel", "Новых заказов для загрузки нет.\n\n" + "\n".join(details))
                return

            message_lines = [
                "Проверка Excel завершена.",
                "",
                f"Файлов: {preview.get('files_count', 0)}",
                f"Строк в файлах: {preview.get('source_rows_count', 0)}",
                f"Новых позиций после объединения: {len(new_records)}",
                f"Клиентов: {preview.get('clients_count', 0)}",
                f"Товаров: {preview.get('products_count', 0)}",
                f"ШТ всего: {preview.get('quantity_count', 0)}",
                f"Блоков к сканированию: {preview.get('blocks_count', 0)}",
                f"Адресов получено из координат: {preview.get('geocoded_count', 0)}",
                f"Координат без адреса: {preview.get('geocode_failed_count', 0)}",
                f"Повторных позиций пропущено: {len(duplicate_records)}",
            ]
            if errors:
                message_lines.extend(["", "Ошибки в отдельных файлах:", "\n".join(errors[:5])])
            if warnings:
                message_lines.extend(["", "Предупреждения:", "\n".join(warnings[:5])])
            message_lines.extend(["", "Загрузить новые позиции в Google Sheets?"])

            if not messagebox.askyesno("Подтверждение импорта", "\n".join(message_lines)):
                self.status_var.set("Импорт отменён")
                return

            self.commit_excel_import(new_records)

        def on_error(exc):
            self.show_critical_error("Не удалось проверить Excel-файлы", exc)

        def on_finally():
            self.import_btn.config(state="normal")
            self.refresh_btn.config(state="normal")
            self.clear_busy()

        self.run_background(
            "Не удалось проверить Excel-файлы",
            work,
            on_success=on_success,
            on_error=on_error,
            on_finally=on_finally
        )

    def commit_excel_import(self, records):
        self.set_busy("⏳ Загружаю заказы в Google Sheets...")
        self.import_btn.config(state="disabled")
        self.refresh_btn.config(state="disabled")

        def work():
            result = append_import_records(records)
            loaded = fetch_sheet_data_with_sync()
            return result, loaded

        def on_success(result):
            import_result, loaded = result
            self.product_catalog = load_product_catalog()
            self.apply_loaded_data(loaded, show_empty_warning=False)
            self.reset_current_selection()
            self.refresh_legal_list()
            messagebox.showinfo(
                "Импорт завершён",
                f"Загружено позиций: {import_result.get('imported', 0)}\n"
                f"Повторно пропущено: {import_result.get('duplicates', 0)}"
            )
            imported_sources = sorted({record.get("Источник файла", "") for record in records if record.get("Источник файла")})
            imported_blocks = sum(parse_int_value(record.get("Кол-во блок")) for record in records)
            if imported_sources:
                self.send_telegram_alert_async(
                    f"{APP_NAME}: импортирован документ\n\n"
                    f"Документы: {', '.join(imported_sources[:5])}\n"
                    f"Позиций загружено: {import_result.get('imported', 0)}\n"
                    f"План КИЗ: {imported_blocks}\n\n"
                    "Документ доступен в разделе «Документы по импорту».",
                    with_keyboard=True,
                )
            self.status_var.set("✅ Excel-заказы загружены")
            self.status_label.config(bg=BG_MAIN, fg=FG_MUTED)

        def on_error(exc):
            self.show_critical_error("Не удалось загрузить Excel-заказы", exc)

        def on_finally():
            self.import_btn.config(state="normal")
            self.refresh_btn.config(state="normal")
            self.clear_busy()

        self.run_background(
            "Не удалось загрузить Excel-заказы",
            work,
            on_success=on_success,
            on_error=on_error,
            on_finally=on_finally
        )

    def show_product_catalog(self):
        if not self.ensure_update_allowed():
            return

        catalog = load_product_catalog()
        dialog = tk.Toplevel(self)
        dialog.title("Справочник товаров")
        dialog.configure(bg=BG_MAIN)
        dialog.geometry("720x480")
        dialog.transient(self)
        dialog.grab_set()

        container = tk.Frame(dialog, bg=BG_MAIN, padx=16, pady=16)
        container.pack(fill="both", expand=True)

        left = tk.Frame(container, bg=BG_CARD, bd=1, highlightbackground=BORDER)
        left.pack(side="left", fill="both", expand=True, padx=(0, 12))

        tk.Label(left, text="Товары", bg=BG_CARD, fg=ACCENT, font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=(12, 8))
        product_list = tk.Listbox(left, bg=BG_CARD, fg=FG_TEXT, relief="flat", font=("Segoe UI", 10))
        product_list.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        right = tk.Frame(container, bg=BG_CARD, bd=1, highlightbackground=BORDER)
        right.pack(side="right", fill="both", expand=True)

        tk.Label(right, text="Карточка товара", bg=BG_CARD, fg=ACCENT, font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=(12, 8))

        name_var = tk.StringVar()
        pieces_var = tk.StringVar(value=str(DEFAULT_PIECES_PER_BLOCK))
        requires_var = tk.BooleanVar(value=True)
        selected_key = {"value": None}

        def field(label, variable):
            row = tk.Frame(right, bg=BG_CARD)
            row.pack(fill="x", padx=12, pady=5)
            tk.Label(row, text=label, bg=BG_CARD, fg=FG_MUTED, font=("Segoe UI", 10), width=18, anchor="w").pack(side="left")
            entry = tk.Entry(
                row,
                textvariable=variable,
                bg=BG_MAIN,
                fg=FG_TEXT,
                relief="flat",
                bd=0,
                font=("Segoe UI", 10),
                highlightbackground=BORDER,
                highlightcolor=ACCENT,
                highlightthickness=1,
                insertbackground=FG_TEXT,
            )
            entry.pack(side="left", fill="x", expand=True)
            return entry

        field("Название", name_var)
        field("ШТ в блоке", pieces_var)
        tk.Checkbutton(
            right,
            text="Нужен КИЗ",
            variable=requires_var,
            bg=BG_CARD,
            fg=FG_TEXT,
            activebackground=BG_CARD,
            font=("Segoe UI", 10)
        ).pack(anchor="w", padx=12, pady=(4, 8))

        def catalog_items():
            return sorted(catalog.items(), key=lambda item: normalize_lookup_text(item[1].get("name") or item[0]))

        def refresh_list():
            product_list.delete(0, tk.END)
            for _, item in catalog_items():
                product_list.insert(tk.END, f"{item.get('name', '')} | {parse_int_value(item.get('pieces_per_block')) or DEFAULT_PIECES_PER_BLOCK} шт.")

        def on_select(_event=None):
            selection = product_list.curselection()
            if not selection:
                return
            key, item = catalog_items()[selection[0]]
            selected_key["value"] = key
            name_var.set(item.get("name", ""))
            pieces_var.set(str(parse_int_value(item.get("pieces_per_block")) or DEFAULT_PIECES_PER_BLOCK))
            requires_var.set(bool(item.get("requires_kiz", True)))

        def save_current():
            name = normalize_text(name_var.get())
            pieces = parse_int_value(pieces_var.get())
            if not name:
                self.show_error("Укажите название товара")
                return
            if pieces <= 0:
                self.show_error("ШТ в блоке должно быть больше нуля")
                return

            old_key = selected_key.get("value")
            new_key = product_catalog_key(name)
            if old_key and old_key != new_key:
                catalog.pop(old_key, None)
            catalog[new_key] = {
                "name": name,
                "pieces_per_block": pieces,
                "requires_kiz": bool(requires_var.get()),
            }
            selected_key["value"] = new_key
            save_product_catalog(catalog)
            self.product_catalog = catalog
            refresh_list()
            self.status_var.set("✅ Справочник товаров сохранён")

        def new_product():
            selected_key["value"] = None
            name_var.set("")
            pieces_var.set(str(DEFAULT_PIECES_PER_BLOCK))
            requires_var.set(True)

        def delete_product():
            key = selected_key.get("value")
            if not key:
                return
            if messagebox.askyesno("Удалить товар?", "Удалить выбранный товар из справочника?"):
                catalog.pop(key, None)
                save_product_catalog(catalog)
                new_product()
                refresh_list()

        product_list.bind("<<ListboxSelect>>", on_select)
        refresh_list()

        actions = tk.Frame(right, bg=BG_CARD)
        actions.pack(fill="x", padx=12, pady=(12, 0))
        AppButton(actions, text="СОХРАНИТЬ", bg=SUCCESS, fg="white", font=("Segoe UI", 9, "bold"), relief="flat", command=save_current).pack(side="left", fill="x", expand=True, padx=(0, 6))
        AppButton(actions, text="НОВЫЙ", bg=INFO, fg="white", font=("Segoe UI", 9, "bold"), relief="flat", command=new_product).pack(side="left", fill="x", expand=True, padx=(0, 6))
        AppButton(actions, text="УДАЛИТЬ", bg=DANGER, fg="white", font=("Segoe UI", 9, "bold"), relief="flat", command=delete_product).pack(side="left", fill="x", expand=True)

        close_frame = tk.Frame(right, bg=BG_CARD)
        close_frame.pack(fill="x", padx=12, pady=(16, 12))
        AppButton(close_frame, text="ЗАКРЫТЬ", bg=FG_MUTED, fg="white", font=("Segoe UI", 9, "bold"), relief="flat", command=dialog.destroy).pack(side="right")

    def show_control_panel(self):
        if not self.ensure_update_allowed():
            return

        if self.operation_in_progress:
            self.show_error("Дождитесь завершения текущей операции")
            return

        self.set_busy("⏳ Собираю контрольную панель...")
        self.control_btn.config(state="disabled")

        def work():
            sheet = self.sheet
            if not sheet:
                client = get_google_client()
                sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)
            return build_control_panel_stats_from_gsheet(sheet)

        def on_success(stats):
            dialog = tk.Toplevel(self)
            dialog.title("Панель контроля")
            dialog.configure(bg=BG_MAIN)
            dialog.geometry("620x560")
            dialog.transient(self)

            container = tk.Frame(dialog, bg=BG_CARD, padx=18, pady=16)
            container.pack(fill="both", expand=True, padx=16, pady=16)

            tk.Label(container, text="Панель контроля за день", bg=BG_CARD, fg=ACCENT, font=("Segoe UI", 14, "bold")).pack(anchor="w", pady=(0, 12))
            rows = [
                ("Заказов по клиенту/адресу", stats.get("groups", 0)),
                ("Активных заказов", stats.get("active_groups", 0)),
                ("Завершённых заказов", stats.get("completed_groups", 0)),
                ("Позиций всего", stats.get("positions", 0)),
                ("Новые позиции", stats.get("new_positions", 0)),
                ("В работе", stats.get("in_progress_positions", 0)),
                ("Завершённые позиции", stats.get("completed_positions", 0)),
                ("План блоков", stats.get("plan_blocks", 0)),
                ("Отсканировано блоков", stats.get("scanned_blocks", 0)),
                ("Осталось блоков", stats.get("remaining_blocks", 0)),
                ("Очередь записи", stats.get("pending_saves", 0)),
                ("Очередь печати", stats.get("pending_prints", 0)),
            ]

            for label, value in rows:
                row = tk.Frame(container, bg=BG_CARD)
                row.pack(fill="x", pady=2)
                tk.Label(row, text=f"{label}:", bg=BG_CARD, fg=FG_MUTED, width=24, anchor="w", font=("Segoe UI", 10)).pack(side="left")
                tk.Label(row, text=str(value), bg=BG_CARD, fg=FG_TEXT, anchor="w", font=("Segoe UI", 10, "bold")).pack(side="left")

            payments = stats.get("payments", {})
            tk.Label(container, text="Оплата", bg=BG_CARD, fg=ACCENT, font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(14, 6))
            payment_text = f"Терминал: {payments.get('terminal', 0)} | Перечисление: {payments.get('transfer', 0)} | Не распознано: {payments.get('unknown', 0)}"
            tk.Label(container, text=payment_text, bg=BG_CARD, fg=FG_TEXT, font=("Segoe UI", 10)).pack(anchor="w")

            tk.Label(container, text="Товары по плану", bg=BG_CARD, fg=ACCENT, font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(14, 6))
            products_text = "\n".join([f"{name}: {blocks} блок." for name, blocks in list(stats.get("products", {}).items())[:12]])
            if not products_text:
                products_text = "Нет данных"
            tk.Label(container, text=products_text, bg=BG_CARD, fg=FG_TEXT, font=("Segoe UI", 10), justify="left", wraplength=540).pack(anchor="w")

            AppButton(container, text="ЗАКРЫТЬ", bg=FG_MUTED, fg="white", font=("Segoe UI", 9, "bold"), relief="flat", command=dialog.destroy).pack(anchor="e", pady=(16, 0))

        def on_error(exc):
            self.show_critical_error("Не удалось собрать панель контроля", exc)

        def on_finally():
            self.control_btn.config(state="normal")
            self.clear_busy()

        self.run_background(
            "Не удалось собрать панель контроля",
            work,
            on_success=on_success,
            on_error=on_error,
            on_finally=on_finally
        )

    def select_legal_entity(self):
        if not self.ensure_update_allowed():
            return

        if self.operation_in_progress:
            self.show_error("Дождитесь завершения текущей операции")
            return

        if not self.today_orders:
            messagebox.showwarning("Ошибка", "Нет доступных юридических лиц!")
            return

        selection = self.legal_listbox.curselection()
        if not selection:
            messagebox.showwarning("Ошибка", "Выберите заказ из списка")
            return

        if selection[0] >= len(self.visible_order_groups):
            messagebox.showwarning("Ошибка", "Выбранный заказ не найден в списке")
            return

        selected_group = self.visible_order_groups[selection[0]]
        legal_entity, payment_type, address = selected_group

        self.current_legal_entity = legal_entity
        self.current_group_key = selected_group
        self.current_legal_entity_orders = [
            o for o in self.today_orders
            if order_group_key(o) == selected_group
        ]
        self.current_legal_entity_orders.sort(key=lambda order: parse_int_value(order.get("_row_number")))
        self.current_product_idx = 0
        self.scanned_codes = []
        self.current_legal_entity_products = []

        self.load_current_product()

        self.status_var.set(f"✅ Выбран заказ: {legal_entity} | {payment_type} | {address}")
        self.scan_entry.focus_set()

    def load_current_product(self):
        if self.current_product_idx >= len(self.current_legal_entity_orders):
            return

        self.current_order = self.current_legal_entity_orders[self.current_product_idx]

        plan_blocks = get_plan_blocks(self.current_order)
        pieces_per_block = get_product_rule(self.current_order.get("Товары", ""), self.product_catalog)["pieces_per_block"]

        info_text = f"""🏢 Юр.лицо: {self.current_order.get('Клиент', '')}
👤 Торг.пред: {self.current_order.get('Торговый представитель', '')}
📍 Адрес: {self.current_order.get('Адрес', 'Адрес не указан')}
📦 Товар: {self.current_order.get('Товары', '')}
💳 Тип оплаты: {self.current_order.get('Тип оплаты', '')}
📦 План: {plan_blocks} блоков (1 блок = {pieces_per_block} ШТ)"""

        self.current_info.config(text=info_text)

        total_products = len(self.current_legal_entity_orders)
        self.position_label.config(text=f"Позиция {self.current_product_idx + 1} из {total_products}")

        existing_codes = self.current_order.get("_existing_scanned_codes", [])
        self.scanned_codes = existing_codes.copy()
        self.saved_codes_count = len(existing_codes)
        self.progress_label.config(text=f"{len(self.scanned_codes)} / {plan_blocks}")
        self.next_product_btn.config(state="disabled")
        self.finish_btn.config(state="normal")
        self.undo_btn.config(state="normal")
        self.scan_entry.delete(0, tk.END)
        if existing_codes:
            self.last_code_label.config(text=f"Уже записано в таблице: {len(existing_codes)} кодов")
        else:
            self.last_code_label.config(text="")
        if plan_blocks > 0 and len(self.scanned_codes) >= plan_blocks:
            self.next_product_btn.config(state="normal")
            self.finish_btn.config(state="disabled")
        self.scan_entry.focus_set()

    def on_scan(self, event=None):
        if not self.ensure_update_allowed():
            self.scan_entry.delete(0, tk.END)
            return

        if self.operation_in_progress:
            self.show_error("Дождитесь завершения текущей операции")
            self.scan_entry.delete(0, tk.END)
            return

        if not self.current_order:
            messagebox.showwarning("Ошибка", "Сначала выберите заказ")
            return

        code = self.scan_entry.get().strip()
        if not code:
            return

        is_valid, error_msg = self.validate_code(code)
        if not is_valid:
            self.show_error(error_msg)
            self.scan_entry.delete(0, tk.END)
            return

        plan_blocks = get_plan_blocks(self.current_order)
        if plan_blocks <= 0:
            self.show_error("В заказе не указано корректное 'Кол-во блок'")
            self.scan_entry.delete(0, tk.END)
            return

        if len(self.scanned_codes) >= plan_blocks:
            self.show_error(f"План выполнен! Нельзя сканировать больше {plan_blocks} блоков")
            self.scan_entry.delete(0, tk.END)
            return

        if code in self.scanned_codes:
            self.show_error("Код уже отсканирован в этой позиции")
            self.scan_entry.delete(0, tk.END)
            return

        if code in self.all_existing_codes:
            self.show_error(f"Код {code[:20]}... уже существует в Google Sheets!")
            self.log_duplicate_code_async(code)
            self.scan_entry.delete(0, tk.END)
            return

        for completed in self.completed_orders:
            if code in completed.get("Коды", []):
                self.show_error("Код уже использован в другом задании сегодня")
                self.scan_entry.delete(0, tk.END)
                return

        if not write_scan_backup("scan", self.current_order, code=code, codes=self.scanned_codes + [code]):
            self.show_error("Не удалось сохранить локальный backup. Код не принят")
            self.scan_entry.delete(0, tk.END)
            return

        self.scanned_codes.append(code)
        self.all_existing_codes.add(code)
        scanned_count = len(self.scanned_codes)

        self.progress_label.config(text=f"{scanned_count} / {plan_blocks}")
        self.last_code_label.config(text=f"Последний код: {code[:40]}...")
        self.status_var.set(f"✅ Отсканирован код ({scanned_count}/{plan_blocks})")
        self.status_label.config(bg=BG_MAIN, fg=FG_MUTED)
        self.scan_entry.delete(0, tk.END)

        if scanned_count >= plan_blocks:
            self.status_var.set(f"🎯 Позиция выполнена! Нажмите 'Следующая позиция'")
            self.next_product_btn.config(state="normal")
            self.finish_btn.config(state="disabled")

        self.scan_entry.focus_set()

    def next_product(self):
        if not self.ensure_update_allowed():
            return

        if self.operation_in_progress:
            self.show_error("Дождитесь завершения текущей операции")
            return

        if not self.current_order:
            return

        plan_blocks = get_plan_blocks(self.current_order)

        scanned_count = len(self.scanned_codes)

        if scanned_count != plan_blocks:
            self.show_error(f"Отсканировано {scanned_count} из {plan_blocks} блоков. Завершите позицию!")
            return

        order = self.current_order
        scanned_codes = self.scanned_codes.copy()
        pieces_per_block = get_product_rule(order.get("Товары", ""), self.product_catalog)["pieces_per_block"]
        self.set_busy("⏳ Сохраняю КИЗы в Google Sheets...")
        self.next_product_btn.config(state="disabled")
        self.finish_btn.config(state="disabled")

        def work():
            ok = False
            message = "Нет подключения к Google Sheets"
            if self.sheet:
                ok, message = update_scanned_codes_to_gsheet(self.sheet, order, scanned_codes)

            if not ok:
                if not is_retryable_save_error(message):
                    raise RuntimeError(message)
                add_pending_save(order, scanned_codes, message)
                if not write_scan_backup("position_queued", order, codes=scanned_codes):
                    raise RuntimeError("Google Sheets недоступен, и локальная очередь записи не создана")
                return {"queued": True, "message": message}

            if not write_scan_backup("position_saved", order, codes=scanned_codes):
                raise RuntimeError("Коды записаны в Google Sheets, но локальный backup позиции не создан")
            return {"queued": False, "message": message}

        def on_success(result):
            product_result = {
                "Клиент": order.get('Клиент', ''),
                "Адрес": order.get('Адрес', ''),
                "Торговый представитель": order.get('Торговый представитель', ''),
                "Товары": order.get('Товары', ''),
                "Тип оплаты": order.get('Тип оплаты', ''),
                "Кол-во ШТ в блоке": pieces_per_block,
                "План": plan_blocks,
                "Отсканировано": scanned_count,
                "Коды": scanned_codes.copy()
            }
            self.current_legal_entity_products.append(product_result)
            order["Отсканированные коды"] = "\n".join(scanned_codes)
            order[STATUS_COLUMN] = get_order_status(order)
            order["_existing_scanned_codes"] = scanned_codes.copy()

            completed_result = product_result.copy()
            completed_result["План блоков"] = plan_blocks
            self.completed_orders.append(completed_result)

            self.current_product_idx += 1
            self.clear_busy()

            if self.current_product_idx < len(self.current_legal_entity_orders):
                self.load_current_product()
                if result.get("queued"):
                    self.status_var.set("⚠️ Позиция сохранена локально, отправится при обновлении")
                else:
                    self.status_var.set("✅ Позиция сохранена")
                self.status_label.config(bg=BG_MAIN, fg=FG_MUTED)
            else:
                self.finish_legal_entity(from_next_product=True)
            self.update_stats_display()

        def on_error(exc):
            self.show_critical_error("КИЗы не записаны", exc)
            self.next_product_btn.config(state="normal")
            self.clear_busy()

        self.run_background(
            "Не удалось сохранить позицию",
            work,
            on_success=on_success,
            on_error=on_error
        )

    def finish_legal_entity(self, from_next_product=False):
        if not self.ensure_update_allowed():
            return

        if self.operation_in_progress:
            self.show_error("Дождитесь завершения текущей операции")
            return

        if not self.current_legal_entity:
            return

        if self.current_product_idx < len(self.current_legal_entity_orders):
            self.show_error("Сначала завершите все позиции по заказу!")
            return

        if not self.current_legal_entity_products:
            self.show_error("Нет завершённых позиций по заказу!")
            return

        if not self.confirm_print_settings():
            self.show_error("Печать сводного листа отменена")
            self.finish_btn.config(state="normal")
            return

        group_key = self.current_group_key
        current_products = [product.copy() for product in self.current_legal_entity_products]
        self.set_busy("⏳ Готовлю и печатаю сводный лист...")
        self.finish_btn.config(state="disabled")
        self.next_product_btn.config(state="disabled")

        def work():
            first_product = current_products[0]
            address = first_product.get('Адрес', 'Адрес не указан')
            summary_products = current_products

            if self.sheet:
                sheet_products = build_summary_products_from_gsheet(
                    self.sheet,
                    group_key or order_group_key(first_product)
                )
                if sheet_products:
                    summary_products = sheet_products
                    first_product = summary_products[0]
                    address = first_product.get('Адрес', address)

            pending_print_id = add_pending_print(address, summary_products)

            printed_files = print_summary(address, summary_products)
            if not printed_files:
                raise RuntimeError("Сводочный лист не создан или не отправлен на печать")

            remove_pending_print(pending_print_id)

            if not write_scan_backup(
                "address_finished",
                first_product,
                codes=[code for product in summary_products for code in product.get("Коды", [])]
            ):
                raise RuntimeError("Сводка напечатана, но backup завершения заказа не создан")

            return {
                "first_product": first_product,
                "summary_products": summary_products,
                "finished_group": group_key or order_group_key(first_product),
            }

        def on_success(result):
            self.update_stats_display()

            finished_group = result["finished_group"]
            self.today_orders = [o for o in self.today_orders if order_group_key(o) != finished_group]
            self.refresh_legal_list()

            self.reset_current_selection()
            self.status_var.set("✅ Заказ завершён! Сводка отправлена на печать")
            self.status_label.config(bg=BG_MAIN, fg=FG_MUTED)

            if self.legal_listbox.size() > 0:
                self.legal_listbox.selection_set(0)

        def on_error(exc):
            self.show_critical_error("Не удалось завершить заказ", exc)
            self.finish_btn.config(state="normal")

        def on_finally():
            self.clear_busy()

        self.run_background(
            "Не удалось завершить заказ",
            work,
            on_success=on_success,
            on_error=on_error,
            on_finally=on_finally
        )

    def confirm_print_settings(self):
        result = {"print": False}
        settings = load_print_settings()
        dialog = tk.Toplevel(self)
        dialog.title("Параметры печати")
        dialog.configure(bg=BG_MAIN)
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        container = tk.Frame(dialog, bg=BG_CARD, padx=24, pady=20)
        container.pack(fill="both", expand=True, padx=16, pady=16)

        tk.Label(
            container,
            text="Печать сводного листа",
            bg=BG_CARD,
            fg=FG_TEXT,
            font=("Segoe UI", 14, "bold")
        ).pack(anchor="w", pady=(0, 12))

        printer_var = tk.StringVar(value=settings.get("printer_name", "Термопринтер"))
        save_var = tk.BooleanVar(value=True)

        printer_row = tk.Frame(container, bg=BG_CARD)
        printer_row.pack(fill="x", pady=3)
        tk.Label(printer_row, text="Принтер:", bg=BG_CARD, fg=FG_MUTED, font=("Segoe UI", 10), width=18, anchor="w").pack(side="left")
        tk.Entry(
            printer_row,
            textvariable=printer_var,
            bg=BG_MAIN,
            fg=FG_TEXT,
            relief="flat",
            bd=0,
            font=("Segoe UI", 10),
            highlightbackground=BORDER,
            highlightcolor=ACCENT,
            highlightthickness=1,
            insertbackground=FG_TEXT,
        ).pack(side="left", fill="x", expand=True)

        rows = [
            ("Размер листа", f"{LABEL_WIDTH_MM} x {LABEL_HEIGHT_MM} мм"),
            ("Масштаб", settings.get("scale", "100%")),
            ("Ориентация", "Квадратная этикетка"),
            ("Разрешение макета", f"{LABEL_DPI} DPI"),
        ]

        for label, value in rows:
            row = tk.Frame(container, bg=BG_CARD)
            row.pack(fill="x", pady=3)
            tk.Label(row, text=label + ":", bg=BG_CARD, fg=FG_MUTED, font=("Segoe UI", 10), width=18, anchor="w").pack(side="left")
            tk.Label(row, text=value, bg=BG_CARD, fg=FG_TEXT, font=("Segoe UI", 10, "bold"), anchor="w").pack(side="left")

        tk.Checkbutton(
            container,
            text="Запомнить параметры",
            variable=save_var,
            bg=BG_CARD,
            fg=FG_TEXT,
            activebackground=BG_CARD,
            font=("Segoe UI", 10)
        ).pack(anchor="w", pady=(8, 0))

        actions = tk.Frame(container, bg=BG_CARD)
        actions.pack(fill="x", pady=(18, 0))

        def confirm():
            result["print"] = True
            if save_var.get():
                save_print_settings({
                    "printer_name": normalize_text(printer_var.get()) or "Термопринтер",
                    "label_width_mm": LABEL_WIDTH_MM,
                    "label_height_mm": LABEL_HEIGHT_MM,
                    "dpi": LABEL_DPI,
                    "scale": "100%",
                })
            dialog.destroy()

        def cancel():
            dialog.destroy()

        AppButton(
            actions,
            text="ПЕЧАТАТЬ",
            bg=SUCCESS,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            padx=18,
            pady=8,
            command=confirm,
            cursor="hand2"
        ).pack(side="right", padx=(8, 0))

        AppButton(
            actions,
            text="ОТМЕНА",
            bg=FG_MUTED,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            padx=18,
            pady=8,
            command=cancel,
            cursor="hand2"
        ).pack(side="right")

        dialog.protocol("WM_DELETE_WINDOW", cancel)
        self.update_idletasks()
        dialog.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - dialog.winfo_width()) // 2
        y = self.winfo_y() + (self.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{x}+{y}")
        self.wait_window(dialog)
        return result["print"]

    def check_pending_prints(self):
        if self.operation_in_progress:
            self.after(1000, self.check_pending_prints)
            return

        pending = load_pending_prints()
        if not pending:
            return

        if not messagebox.askyesno(
            "Непечатанные сводки",
            f"Найдено непечатанных сводок: {len(pending)}.\n\nНапечатать сейчас?"
        ):
            return

        if not self.confirm_print_settings():
            return

        self.set_busy("⏳ Печатаю сводки из очереди...")

        def work():
            printed_count = 0
            for item in pending[:]:
                printed_files = print_summary(item.get("address", "Адрес не указан"), item.get("products", []))
                if not printed_files:
                    raise RuntimeError("Не удалось напечатать сводку из очереди")
                remove_pending_print(item.get("id"))
                printed_count += 1
            return printed_count

        def on_success(printed_count):
            self.status_var.set(f"✅ Напечатано сводок из очереди: {printed_count}")
            self.status_label.config(bg=BG_MAIN, fg=FG_MUTED)

        def on_error(exc):
            self.show_critical_error("Не удалось напечатать сводки из очереди", exc)

        def on_finally():
            self.clear_busy()

        self.run_background(
            "Не удалось напечатать сводки из очереди",
            work,
            on_success=on_success,
            on_error=on_error,
            on_finally=on_finally
        )

    def update_stats_display(self):
        if not hasattr(self, "completed_count_label"):
            return
        completed = len(self.completed_orders)
        total_blocks = sum(o.get("Отсканировано", 0) for o in self.completed_orders)
        self.completed_count_label.config(text=f"Выполнено: {completed}")
        self.total_blocks_label.config(text=f"Блоков: {total_blocks}")
        active_groups = len({order_group_key(order) for order in self.today_orders})
        pending_saves = len(load_pending_saves())
        self.active_orders_label.config(text=f"Активных заказов: {active_groups}")
        self.pending_saves_label.config(text=f"Очередь записи: {pending_saves}")

    def end_day(self):
        if not self.ensure_update_allowed():
            return

        if self.operation_in_progress:
            self.show_error("Дождитесь завершения текущей операции")
            return

        if self.current_legal_entity:
            if not messagebox.askyesno("Внимание", "У вас есть незавершённый заказ!\n\nЗавершить день без сохранения текущего заказа?"):
                return

        self.set_busy("⏳ Формирую Excel-отчёт за день...")
        self.report_btn.config(state="disabled")

        def work():
            sheet = self.sheet
            if not sheet:
                client = get_google_client()
                sheet = client.open_by_key(SPREADSHEET_ID).worksheet(SHEET_NAME)

            result = create_day_report_excel(sheet)
            result["sheet"] = sheet
            return result

        def on_success(result):
            self.sheet = result.get("sheet") or self.sheet
            if result.get("empty"):
                messagebox.showwarning("Нет данных", "За сегодня в Google Sheets нет отсканированных КИЗов для отчёта")
                return

            total_report_rows = result["total_report_rows"]
            self.send_telegram_documents_async(
                collect_operational_documents(include_report=result.get("filename"))
            )
            messagebox.showinfo("Отчёт сохранён",
                f"📊 Отчёт сохранён: {result['filename']}\n\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"✅ Строк КИЗов: {total_report_rows}\n"
                f"📦 Блоков: {total_report_rows}\n"
                f"🔢 Кодов: {total_report_rows}\n"
                f"├─ Терминал: {result['terminal_count']} кодов\n"
                f"├─ Перечисление: {result['transfer_count']} кодов\n"
                f"└─ Не распознано: {result['unknown_count']} кодов\n"
                f"━━━━━━━━━━━━━━━━━━━━")

            self.on_close()

        def on_error(exc):
            if isinstance(exc, ImportError):
                self.show_critical_error("Не установлены зависимости для Excel-отчёта", "Установите pandas и openpyxl:\npip install pandas openpyxl")
            else:
                self.show_critical_error("Не удалось сохранить Excel-отчёт", exc)

        def on_finally():
            try:
                if self.winfo_exists():
                    self.report_btn.config(state="normal")
                    self.clear_busy()
            except tk.TclError:
                pass

        self.run_background(
            "Не удалось сохранить Excel-отчёт",
            work,
            on_success=on_success,
            on_error=on_error,
            on_finally=on_finally
        )

    def on_close(self):
        if self.current_order and len(self.scanned_codes) > self.saved_codes_count:
            if not messagebox.askyesno(
                "Закрыть программу?",
                "Есть несохранённые сканы по текущей позиции.\n\nЗакрыть программу без завершения позиции?"
            ):
                return
        self.destroy()

if __name__ == "__main__":
    if maybe_rename_windows_executable():
        sys.exit(0)

    if not os.path.exists(CREDENTIALS_FILE):
        messagebox.showerror("Ошибка",
            f"Файл {CREDENTIALS_FILE} не найден!\n\n"
            f"Положите файл с учётными данными Google Sheets в папку с программой")
    else:
        app = ScanningApp()
        app.mainloop()
