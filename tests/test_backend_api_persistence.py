import unittest
import uuid
from io import BytesIO
from datetime import date

import openpyxl
from fastapi.testclient import TestClient
from sqlalchemy import create_engine, select
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend.app.db import get_db
from backend.app.main import app, require_service_token
from backend.app.models import AuditLog, Base, Order, OrderItem, ScanCode


class BackendApiPersistenceTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite+pysqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=self.engine)

        def override_get_db():
            db = self.SessionLocal()
            try:
                yield db
            finally:
                db.close()

        app.dependency_overrides[get_db] = override_get_db
        app.dependency_overrides[require_service_token] = lambda: None
        self.client = TestClient(app)

    def tearDown(self):
        app.dependency_overrides.clear()
        Base.metadata.drop_all(self.engine)
        self.engine.dispose()

    def seed_order(self, *, status="not_completed", quantity_blocks=2, scanned_blocks=0, item_status="not_completed"):
        with self.SessionLocal() as db:
            order = Order(
                payment_type="cash",
                client="Test Client",
                address="Test Address",
                representative="Test Rep",
                order_date=date(2026, 5, 30),
                status=status,
                raw_payload={"source": "test"},
            )
            item = OrderItem(
                order=order,
                product="Test Product",
                quantity_pieces=20,
                quantity_blocks=quantity_blocks,
                pieces_per_block=10,
                scanned_blocks=scanned_blocks,
                status=item_status,
                raw_payload={"source": "test"},
            )
            db.add_all([order, item])
            db.commit()
            return str(order.id), str(item.id)

    def test_active_orders_returns_uncompleted_orders_with_items(self):
        active_order_id, _ = self.seed_order()
        self.seed_order(status="completed", item_status="completed")

        response = self.client.get("/api/v1/orders/active")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(len(payload), 1)
        self.assertEqual(payload[0]["id"], active_order_id)
        self.assertEqual(payload[0]["status"], "not_completed")
        self.assertEqual(payload[0]["items"][0]["product"], "Test Product")

    def test_scan_creates_code_increments_item_and_rejects_duplicate(self):
        _, item_id = self.seed_order()

        response = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "  010123456789  ", "workstation_id": "pc-1"},
        )

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertEqual(payload["code"], "010123456789")
        self.assertEqual(payload["scanned_blocks"], 1)
        self.assertEqual(payload["item_status"], "not_completed")

        duplicate = self.client.post(
            "/api/v1/scans",
            json={"order_item_id": item_id, "code": "010123456789", "workstation_id": "pc-2"},
        )
        self.assertEqual(duplicate.status_code, 409)
        self.assertEqual(duplicate.json()["detail"], "Code already scanned")

        with self.SessionLocal() as db:
            self.assertEqual(len(db.execute(select(ScanCode)).scalars().all()), 1)
            item = db.get(OrderItem, uuid.UUID(item_id))
            self.assertEqual(item.scanned_blocks, 1)

    def test_complete_order_requires_required_blocks_and_closes_order(self):
        order_id, item_id = self.seed_order()

        too_early = self.client.post(f"/api/v1/orders/{order_id}/complete")

        self.assertEqual(too_early.status_code, 409)
        self.assertEqual(too_early.json()["detail"]["message"], "Order has incomplete required items")

        for code in ["010000000001", "010000000002"]:
            scan = self.client.post("/api/v1/scans", json={"order_item_id": item_id, "code": code})
            self.assertEqual(scan.status_code, 201)

        completed = self.client.post(f"/api/v1/orders/{order_id}/complete")

        self.assertEqual(completed.status_code, 200)
        self.assertEqual(completed.json()["status"], "completed")
        self.assertEqual(completed.json()["items"][0]["status"], "completed")

        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        self.assertEqual(active.json(), [])

        with self.SessionLocal() as db:
            actions = [row.action for row in db.execute(select(AuditLog)).scalars().all()]
            self.assertEqual(actions.count("scan_code_created"), 2)
            self.assertIn("order_completed", actions)

    def test_import_creates_grouped_order_items_and_history(self):
        rows = [
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "cash",
                "Клиент": "Import Client",
                "Адрес": "Import Address",
                "Торговый представитель": "Import Rep",
                "Товары": "Product One",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "ID заказа": "source-order-1",
                "ID импорта": "import-row-1",
            },
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "cash",
                "Клиент": "Import Client",
                "Адрес": "Import Address",
                "Торговый представитель": "Import Rep",
                "Товары": "Product Two",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "ID заказа": "source-order-2",
                "ID импорта": "import-row-2",
            },
        ]

        response = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})

        self.assertEqual(response.status_code, 201)
        payload = response.json()
        self.assertEqual(payload["status"], "completed")
        self.assertEqual(payload["orders_created"], 1)
        self.assertEqual(payload["items_created"], 2)
        self.assertEqual(payload["duplicate_rows"], 0)

        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        active_payload = active.json()
        self.assertEqual(len(active_payload), 1)
        self.assertEqual(active_payload[0]["client"], "Import Client")
        self.assertEqual(len(active_payload[0]["items"]), 2)
        self.assertEqual(active_payload[0]["items"][0]["scan_codes"], [])

        history = self.client.get("/api/v1/imports")
        self.assertEqual(history.status_code, 200)
        self.assertEqual(len(history.json()), 1)
        self.assertEqual(history.json()[0]["rows_imported"], 2)

    def test_import_stores_coordinates_blocks_and_prices(self):
        rows = [
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "Терминал",
                "Клиент": "Price Client",
                "Адрес": "Tashkent Address",
                "Координаты": "41.31, 69.27",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "200",
                "Сумма позиции": "4800000",
                "ID заказа": "price-source-order",
            },
        ]

        response = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})

        self.assertEqual(response.status_code, 201)
        active = self.client.get("/api/v1/orders/active")
        self.assertEqual(active.status_code, 200)
        order = active.json()[0]
        self.assertEqual(order["coordinates"], "41.31, 69.27")
        self.assertEqual(order["items"][0]["quantity_pieces"], 200)
        self.assertEqual(order["items"][0]["quantity_blocks"], 20)
        self.assertEqual(order["items"][0]["block_price"], 240000)
        self.assertEqual(order["items"][0]["line_total"], 4_800_000)

    def test_import_skips_duplicate_items_and_reports_invalid_rows(self):
        valid_row = {
            "Дата отгрузки": "2026-05-30",
            "Тип оплаты": "cash",
            "Клиент": "Duplicate Client",
            "Адрес": "Duplicate Address",
            "Товары": "Duplicate Product",
            "Кол-во ШТ": 20,
            "Кол-во блок": 2,
            "ID заказа": "duplicate-source-order",
        }
        first = self.client.post("/api/v1/imports", json={"source": "excel", "rows": [valid_row]})
        second = self.client.post("/api/v1/imports", json={"source": "excel", "rows": [valid_row, {"Клиент": "Broken"}]})

        self.assertEqual(first.status_code, 201)
        self.assertEqual(second.status_code, 201)
        payload = second.json()
        self.assertEqual(payload["items_created"], 0)
        self.assertEqual(payload["duplicate_rows"], 1)
        self.assertEqual(payload["invalid_rows"], 1)
        self.assertEqual(payload["status"], "failed")

        with self.SessionLocal() as db:
            self.assertEqual(len(db.execute(select(Order)).scalars().all()), 1)
            self.assertEqual(len(db.execute(select(OrderItem)).scalars().all()), 1)

    def test_day_report_summarizes_orders_scans_and_payment_groups(self):
        rows = [
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "Терминал",
                "Клиент": "Report Client",
                "Адрес": "Report Address",
                "Торговый представитель": "Report Rep",
                "Товары": "Report Product One",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "Номер заявки SkladBot": "WR-100",
                "ID заказа": "report-source-order-1",
            },
            {
                "Дата отгрузки": "30.05.2026",
                "Тип оплаты": "Терминал",
                "Клиент": "Report Client",
                "Адрес": "Report Address",
                "Торговый представитель": "Report Rep",
                "Товары": "Report Product Two",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "Номер заявки SkladBot": "WR-100",
                "ID заказа": "report-source-order-2",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        active = self.client.get("/api/v1/orders/active").json()
        order_id = active[0]["id"]
        item_ids = {item["product"]: item["id"] for item in active[0]["items"]}

        scans = [
            ("Report Product One", "010000000101"),
            ("Report Product One", "010000000102"),
            ("Report Product Two", "010000000201"),
        ]
        for product, code in scans:
            response = self.client.post(
                "/api/v1/scans",
                json={
                    "order_item_id": item_ids[product],
                    "code": code,
                    "scanned_at": "2026-05-30T12:00:00+00:00",
                },
            )
            self.assertEqual(response.status_code, 201)

        active_after_scans = self.client.get("/api/v1/orders/active")
        self.assertEqual(active_after_scans.status_code, 200)
        active_item = active_after_scans.json()[0]["items"][0]
        self.assertTrue(active_item["scan_codes"])

        completed = self.client.post(f"/api/v1/orders/{order_id}/complete")
        self.assertEqual(completed.status_code, 200)

        report = self.client.get("/api/v1/reports/day?report_date=2026-05-30")

        self.assertEqual(report.status_code, 200)
        payload = report.json()
        self.assertEqual(payload["report_date"], "2026-05-30")
        self.assertEqual(payload["source"], "postgres")
        self.assertEqual(payload["totals"]["orders"], 1)
        self.assertEqual(payload["totals"]["completed_orders"], 1)
        self.assertEqual(payload["totals"]["active_orders"], 0)
        self.assertEqual(payload["totals"]["items"], 2)
        self.assertEqual(payload["totals"]["completed_items"], 2)
        self.assertEqual(payload["totals"]["planned_blocks"], 3)
        self.assertEqual(payload["totals"]["scanned_blocks"], 3)
        self.assertEqual(payload["totals"]["scanned_today"], 3)
        self.assertEqual(payload["totals"]["remaining_blocks"], 0)
        self.assertEqual(payload["totals"]["scan_codes"], 3)
        self.assertEqual(payload["payment_groups"][0]["payment_group"], "terminal")
        self.assertEqual(payload["payment_groups"][0]["orders"], 1)
        self.assertEqual(payload["orders"][0]["skladbot_request_number"], "WR-100")

    def test_logistics_report_uses_shipment_date_coordinates_and_prices(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Logistics Client",
                "Адрес": "Tashkent Address",
                "Координаты": "41.31, 69.27",
                "Торговый представитель": "Rep One",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "200",
                "Кол-во блок": "20",
                "Сумма позиции": "4800000",
                "ID заказа": "logistics-source-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        dates = self.client.get("/api/v1/logistics/dates")
        self.assertEqual(dates.status_code, 200)
        self.assertEqual(dates.json(), ["2026-05-30"])

        report = self.client.get("/api/v1/logistics/report?shipment_date=2026-05-30")
        self.assertEqual(report.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(report.content), data_only=True)
        sheet = workbook["Заявки"]

        self.assertEqual(sheet["C2"].value, "Logistics Client")
        self.assertEqual(sheet["G2"].value, "41.31,69.27")
        self.assertEqual(sheet["J2"].value, "30.05.2026")
        self.assertEqual(sheet["R2"].value, "Chapman Brown OP 20")
        self.assertEqual(sheet["S2"].value, 200)
        self.assertEqual(sheet["V2"].value, 24000)
        self.assertEqual(sheet["W2"].value, 4_800_000)
        self.assertEqual(sheet["AE2"].value, "41.31,69.27")
        self.assertEqual(sheet["AF2"].value, "41.31")
        self.assertEqual(sheet["AG2"].value, "69.27")
        workbook.close()

    def test_logistics_report_requires_coordinates(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "No Coordinates Client",
                "Адрес": "Tashkent Address",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "ID заказа": "no-coordinates-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        report = self.client.get("/api/v1/logistics/report?shipment_date=2026-05-30")

        self.assertEqual(report.status_code, 409)
        self.assertIn("Missing coordinates", report.json()["detail"])

    def test_logistics_report_normalizes_three_part_coordinates(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "Coordinates Client",
                "Адрес": "Tashkent Address",
                "Координаты": "41.214609,69.223027,15",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "ID заказа": "coordinates-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        report = self.client.get("/api/v1/logistics/report?shipment_date=2026-05-30")
        self.assertEqual(report.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(report.content), data_only=True)
        sheet = workbook["Заявки"]

        self.assertEqual(sheet["AE2"].value, "41.214609,69.223027")
        self.assertEqual(sheet["AF2"].value, "41.214609")
        self.assertEqual(sheet["AG2"].value, "69.223027")
        workbook.close()

    def test_kiz_source_file_report_lists_only_completed_source_files(self):
        rows = [
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Терминал",
                "Клиент": "KIZ Client",
                "Адрес": "KIZ Address",
                "Координаты": "41.31, 69.27",
                "Товары": "Chapman Brown OP 20",
                "Кол-во ШТ": "20",
                "Кол-во блок": "2",
                "Сумма позиции": "480000",
                "Источник файла": "source-a.xlsx",
                "ID заказа": "kiz-source-order",
            },
            {
                "Дата отгрузки": "2026-05-30",
                "Тип оплаты": "Перечисление",
                "Клиент": "Open Client",
                "Адрес": "Open Address",
                "Товары": "Chapman Gold SSL 20",
                "Кол-во ШТ": "10",
                "Кол-во блок": "1",
                "Источник файла": "source-b.xlsx",
                "ID заказа": "open-source-order",
            },
        ]
        imported = self.client.post("/api/v1/imports", json={"source": "excel", "filename": "orders.xlsx", "rows": rows})
        self.assertEqual(imported.status_code, 201)

        active = self.client.get("/api/v1/orders/active").json()
        kiz_order = next(order for order in active if order["client"] == "KIZ Client")
        item_id = kiz_order["items"][0]["id"]
        for code in ("010000000301", "010000000302"):
            response = self.client.post("/api/v1/scans", json={"order_item_id": item_id, "code": code})
            self.assertEqual(response.status_code, 201)

        source_files = self.client.get("/api/v1/reports/kiz/source-files")
        self.assertEqual(source_files.status_code, 200)
        self.assertEqual([item["source_file"] for item in source_files.json()], ["source-a.xlsx"])

        report = self.client.get("/api/v1/reports/kiz/source-file", params={"source_file": "source-a.xlsx"})
        self.assertEqual(report.status_code, 200)
        workbook = openpyxl.load_workbook(BytesIO(report.content), data_only=True)
        summary = workbook["Сводка"]
        self.assertEqual(summary["C2"].value, "KIZ Client")
        self.assertEqual(summary["G2"].value, 2)
        self.assertEqual(summary["H2"].value, 2)
        self.assertEqual(summary["I2"].value, 480000)
        sheet = workbook["Терминал"]
        self.assertEqual(sheet["C2"].value, "KIZ Client")
        self.assertEqual(sheet["G2"].value, "Chapman Brown OP 20")
        self.assertEqual(sheet["H2"].value, 2)
        self.assertEqual(sheet["I2"].value, "010000000301")
        self.assertEqual(sheet["I3"].value, "010000000302")
        self.assertEqual(sheet["K2"].value, "source-a.xlsx")
        workbook.close()

    def test_day_report_rejects_invalid_report_date(self):
        response = self.client.get("/api/v1/reports/day?report_date=not-a-date")

        self.assertEqual(response.status_code, 422)
        self.assertIn("Invalid report_date", response.json()["detail"])


if __name__ == "__main__":
    unittest.main()
