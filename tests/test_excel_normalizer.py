import importlib
import sys
import tempfile
import types
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from taksklad.geocoding import clean_geocoded_address
from taksklad.excel_normalizer import detect_excel_source, is_summary_row


def import_excel_import():
    try:
        return importlib.import_module("taksklad.excel_import")
    except ModuleNotFoundError as exc:
        if exc.name != "gspread":
            raise

    fake_sheets = types.ModuleType("taksklad.sheets")
    for name in [
        "build_import_record_row",
        "ensure_import_sheet_layout",
        "get_existing_import_keys",
        "get_existing_order_duplicate_keys",
        "get_google_client",
    ]:
        setattr(fake_sheets, name, lambda *args, **kwargs: None)
    sys.modules["taksklad.sheets"] = fake_sheets
    return importlib.import_module("taksklad.excel_import")


class ExcelNormalizerTests(unittest.TestCase):
    def test_clean_geocoded_address_removes_country_prefix(self):
        self.assertEqual(
            clean_geocoded_address("Узбекистан, Ташкент, улица Укчи, 3"),
            "Ташкент, улица Укчи, 3",
        )

    def save_constructor_report(self, path):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Конструктор отчетов"
        worksheet.append(["Статус:  В обработке"])
        worksheet.append(["Дата заказа:  21.05.2026-22.05.2026"])
        worksheet.append(["", "", "", "", "", "Дата заказа", "22.05.2026", "", "ИТОГО", ""])
        worksheet.append([
            "Торговый представитель",
            "Клиент",
            "Координаты клиента",
            "ТМЦ",
            "Тип оплаты",
            "Статус",
            "Количество заказа",
            "Сумма с переоценкой",
            "Количество заказа",
            "Сумма с переоценкой",
        ])
        worksheet.append([
            "ТП1 Суюнбеков Умид Бахрдирович",
            '"BUSINESS MURODOV TRADE" MCHJ (1 филиал)',
            "41.373879,69.322741",
            "Chapman Brown OP 20",
            "Терминал",
            "В обработке",
            20,
            480000,
            20,
            480000,
        ])
        worksheet.append(["ИТОГО", "ИТОГО", "ИТОГО", "ИТОГО", "ИТОГО", "ИТОГО", 20, 480000, 20, 480000])
        workbook.save(path)

    def test_detects_constructor_report_header_and_first_quantity_column(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "Шаблон_отправки_заказов_на_склад_25_05_2026.xlsx"
            self.save_constructor_report(path)

            workbook = load_workbook(path, data_only=True, read_only=True)
            source = detect_excel_source(workbook, str(path))

        self.assertEqual(source["sheet_name"], "Конструктор отчетов")
        self.assertEqual(source["header_row"], 4)
        self.assertEqual(source["columns"]["quantity"], 6)
        self.assertEqual(source["default_date"], "25.05.2026")

    def test_parses_constructor_report_as_import_records(self):
        excel_import = import_excel_import()
        excel_import.reverse_geocode_yandex = lambda coords, cache=None: (f"Адрес {coords}", "")

        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "Шаблон_отправки_заказов_на_склад_25_05_2026.xlsx"
            self.save_constructor_report(path)
            result = excel_import.parse_excel_order_files([str(path)])

        self.assertEqual(result["errors"], [])
        self.assertEqual(result["source_rows_count"], 1)
        self.assertEqual(len(result["records"]), 1)
        record = result["records"][0]
        self.assertEqual(record["Дата отгрузки"], "25.05.2026")
        self.assertEqual(record["Тип оплаты"], "Терминал")
        self.assertEqual(record["Товары"], "Chapman Brown OP 20")
        self.assertEqual(record["Кол-во ШТ"], 20)

    def test_summary_rows_are_skipped(self):
        row = ["ИТОГО", "ИТОГО", "ИТОГО", "ИТОГО", "ИТОГО", "ИТОГО", 20]
        columns = {"client": 1, "payment": 4, "product": 3, "quantity": 6}

        self.assertTrue(is_summary_row(row, columns))


if __name__ == "__main__":
    unittest.main()
