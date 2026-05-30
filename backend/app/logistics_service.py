import re
from io import BytesIO
from re import sub

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .models import Order, OrderItem
from .orders_service import ApiError
from .reports_service import parse_report_date


LOGISTICS_HEADERS = [
    "Тип заявки*",
    "Склад Поставщик*",
    "ФИО или Наименование торговой точки*",
    "E-mail клиента",
    "Номер телефона*",
    "Адрес доставки*",
    "Координаты",
    "Детали адреса",
    "Тип оплаты",
    "Дата доставки*",
    "Доставка С*",
    "Доставка ПО*",
    "Дата забора",
    "Забор С",
    "Забор ПО",
    "Код (товара)",
    "",
    "Наименование Товара",
    "Кол-во",
    "Вес",
    "Объем",
    "Цена",
    "Цена заказа",
    "Доп.информация",
    "Сведения",
    "Категория заказа",
    "Имя отправителя",
    "Номер отправителя",
    "Время загрузки",
    "Время отгрузки",
    "Координаты",
    "Широта (Куда)",
    "Долгота (Куда)",
    "Широта (Откуда)",
    "Долгота (Откуда)",
    "Id заявки",
    "Приоритет",
    "Cтоимость рейса",
    "Теги заявок",
]


def list_logistics_dates(db: Session):
    rows = db.execute(
        select(Order.order_date)
        .where(Order.order_date.is_not(None))
        .distinct()
        .order_by(Order.order_date.asc())
    ).scalars().all()
    return [row.isoformat() for row in rows if row]


def build_logistics_report_xlsx(db: Session, shipment_date: str):
    report_date = parse_report_date(shipment_date)
    orders = db.execute(
        select(Order)
        .options(selectinload(Order.items))
        .where(Order.order_date == report_date)
        .order_by(Order.client.asc(), Order.created_at.asc())
    ).scalars().all()
    if not orders:
        raise ApiError(404, f"No orders for shipment date {report_date.isoformat()}")
    missing_coordinates = [
        order.client
        for order in orders
        if not normalize_coordinates((order.raw_payload or {}).get("coordinates"))
    ]
    if missing_coordinates:
        sample = ", ".join(missing_coordinates[:5])
        raise ApiError(409, f"Missing coordinates for logistics report: {sample}")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Заявки"
    sheet.append(LOGISTICS_HEADERS)
    apply_header_style(sheet)

    for order in orders:
        coordinates = normalize_coordinates((order.raw_payload or {}).get("coordinates"))
        latitude, longitude = split_coordinates(coordinates)
        for item in sorted(order.items, key=lambda value: (value.product, str(value.id))):
            quantity_pieces = item.quantity_pieces or ((item.quantity_blocks or 0) * (item.pieces_per_block or 10))
            line_total = parse_int((item.raw_payload or {}).get("line_total"))
            unit_price = int(line_total / quantity_pieces) if line_total and quantity_pieces else 0
            row = [""] * len(LOGISTICS_HEADERS)
            set_cell(row, 1, "Доставка")
            set_cell(row, 3, order.client)
            set_cell(row, 5, order.representative or "")
            set_cell(row, 7, coordinates)
            set_cell(row, 9, order.payment_type)
            set_cell(row, 10, report_date.strftime("%d.%m.%Y"))
            set_cell(row, 11, "10:00")
            set_cell(row, 12, "18:00")
            set_cell(row, 18, item.product)
            set_cell(row, 19, quantity_pieces)
            set_cell(row, 22, unit_price)
            set_cell(row, 23, line_total)
            set_cell(row, 31, coordinates)
            set_cell(row, 32, latitude)
            set_cell(row, 33, longitude)
            set_cell(row, 36, (order.raw_payload or {}).get("skladbot_request_number") or "")
            sheet.append(row)

    autosize_columns(sheet)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), logistics_report_filename(report_date)


def logistics_report_filename(report_date):
    return f"TakSklad_логистика_{report_date.strftime('%d.%m.%Y')}.xlsx"


def set_cell(row, one_based_index, value):
    row[one_based_index - 1] = value


def normalize_coordinates(value):
    text = str(value or "").strip()
    if not text:
        return ""
    numbers = re.findall(r"-?\d+(?:[.,]\d+)?", text)
    if len(numbers) < 2:
        return ""
    latitude = numbers[0].replace(",", ".")
    longitude = numbers[1].replace(",", ".")
    return f"{latitude},{longitude}"


def split_coordinates(value):
    parts = [part.strip() for part in str(value or "").split(",")]
    if len(parts) < 2:
        return "", ""
    return parts[0], parts[1]


def parse_int(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def apply_header_style(sheet):
    fill = PatternFill("solid", fgColor="F0E68C")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = fill
    sheet.freeze_panes = "A2"


def autosize_columns(sheet):
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 45)


def safe_filename(value):
    return sub(r"[^0-9A-Za-zА-Яа-я_.-]+", "_", str(value or "")).strip("_")
