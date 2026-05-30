from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from .models import Order, OrderItem
from .orders_service import ApiError, COMPLETED_STATUSES
from .reports_service import payment_group


KIZ_REPORT_HEADERS = [
    "Дата отгрузки",
    "Номер заявки SkladBot",
    "Клиент",
    "Адрес",
    "Координаты",
    "Тип оплаты",
    "Товар",
    "Кол-во блок",
    "КИЗ",
    "Цена заказа",
    "Источник файла",
]

KIZ_SUMMARY_HEADERS = [
    "Дата отгрузки",
    "Номер заявки SkladBot",
    "Клиент",
    "Адрес",
    "Координаты",
    "Тип оплаты",
    "План блоков",
    "Отсканировано блоков",
    "Цена заказа",
    "Источник файла",
]


def list_completed_kiz_source_files(db: Session):
    groups = group_items_by_source_file(load_items(db))
    result = []
    for source_file, items in sorted(groups.items()):
        if not source_file or not items:
            continue
        planned_blocks = sum(max(0, item.quantity_blocks or 0) for item in items)
        scanned_blocks = sum(max(0, item.scanned_blocks or 0) for item in items)
        completed = all(item_is_completed(item) for item in items)
        if not completed:
            continue
        dates = sorted({item.order.order_date.isoformat() for item in items if item.order and item.order.order_date})
        result.append({
            "source_file": source_file,
            "dates": dates,
            "items": len(items),
            "planned_blocks": planned_blocks,
            "scanned_blocks": scanned_blocks,
        })
    return result


def build_kiz_source_file_report_xlsx(db: Session, source_file: str):
    source_file = str(source_file or "").strip()
    if not source_file:
        raise ApiError(422, "source_file is required")

    items = [
        item
        for item in load_items(db)
        if str((item.raw_payload or {}).get("source_file") or "").strip() == source_file
    ]
    if not items:
        raise ApiError(404, f"No rows for source file {source_file}")
    if not all(item_is_completed(item) for item in items):
        raise ApiError(409, f"Source file {source_file} is not fully completed")

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    summary_sheet.append(KIZ_SUMMARY_HEADERS)
    apply_header_style(summary_sheet)
    for row in build_summary_rows(items, source_file):
        summary_sheet.append(row)
    autosize_columns(summary_sheet)

    grouped = {}
    for item in items:
        grouped.setdefault(payment_group(item.order.payment_type if item.order else ""), []).append(item)

    for group, group_items in sorted(grouped.items()):
        sheet = workbook.create_sheet(payment_sheet_title(group))
        sheet.append(KIZ_REPORT_HEADERS)
        apply_header_style(sheet)
        for item in sorted(group_items, key=item_sort_key):
            order = item.order
            raw_payload = item.raw_payload or {}
            order_raw = order.raw_payload if order else {}
            codes = [scan.code for scan in sorted(item.scan_codes, key=lambda value: (str(value.scanned_at or ""), str(value.id)))]
            for code in codes:
                sheet.append([
                    order.order_date.strftime("%d.%m.%Y") if order and order.order_date else "",
                    (order_raw or {}).get("skladbot_request_number") or "",
                    order.client if order else "",
                    order.address if order else "",
                    (order_raw or {}).get("coordinates") or "",
                    order.payment_type if order else "",
                    item.product,
                    item.quantity_blocks,
                    code,
                    parse_int(raw_payload.get("line_total")),
                    source_file,
                ])
        autosize_columns(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), kiz_source_file_report_filename(source_file)


def build_summary_rows(items, source_file):
    grouped = {}
    for item in items:
        if not item.order:
            continue
        grouped.setdefault(item.order.id, {"order": item.order, "items": []})["items"].append(item)

    rows = []
    for group in sorted(grouped.values(), key=lambda value: (str(value["order"].order_date or ""), value["order"].client)):
        order = group["order"]
        order_items = group["items"]
        order_raw = order.raw_payload or {}
        planned_blocks = sum(max(0, item.quantity_blocks or 0) for item in order_items)
        scanned_blocks = sum(max(0, item.scanned_blocks or 0) for item in order_items)
        order_total = sum(parse_int((item.raw_payload or {}).get("line_total")) for item in order_items)
        rows.append([
            order.order_date.strftime("%d.%m.%Y") if order.order_date else "",
            order_raw.get("skladbot_request_number") or "",
            order.client,
            order.address,
            order_raw.get("coordinates") or "",
            order.payment_type,
            planned_blocks,
            scanned_blocks,
            order_total,
            source_file,
        ])
    return rows


def load_items(db: Session):
    return db.execute(
        select(OrderItem)
        .options(
            selectinload(OrderItem.order),
            selectinload(OrderItem.scan_codes),
        )
    ).scalars().all()


def group_items_by_source_file(items):
    groups = {}
    for item in items:
        source_file = str((item.raw_payload or {}).get("source_file") or "").strip()
        groups.setdefault(source_file, []).append(item)
    return groups


def item_is_completed(item):
    if item.status in COMPLETED_STATUSES:
        return True
    return (item.quantity_blocks or 0) > 0 and (item.scanned_blocks or 0) >= (item.quantity_blocks or 0)


def item_sort_key(item):
    order = item.order
    return (
        str(order.order_date or "") if order else "",
        order.client if order else "",
        item.product,
        str(item.id),
    )


def payment_sheet_title(group):
    if group == "terminal":
        return "Терминал"
    if group == "transfer":
        return "Перевод"
    return "Неизвестно"


def kiz_source_file_report_filename(source_file):
    safe_name = safe_filename(source_file.rsplit(".", 1)[0])
    return f"TakSklad_КИЗ_{safe_name}.xlsx"


def safe_filename(value):
    allowed = []
    for char in str(value or ""):
        if char.isalnum() or char in "._- ":
            allowed.append(char)
        else:
            allowed.append("_")
    result = "".join(allowed).strip(" ._")
    return result or "file"


def parse_int(value):
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def apply_header_style(sheet):
    fill = PatternFill("solid", fgColor="F0E68C")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="000000")
        cell.fill = fill
    sheet.freeze_panes = "A2"


def autosize_columns(sheet):
    for column_cells in sheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 50)
